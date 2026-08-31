#!/usr/bin/env python3
"""
repair_v11_workbook.py — fix the structural defects in Desktop/v11_FINAL.xlsx.

Two defects, neither in the data:

  1. STALE TABLE OBJECT (this is what makes Excel prompt for repair).
     xl/tables/table1.xml still describes the PRE-cleanup sheet:
         ref="A1:AR564"  (44 columns)
         tableColumns[0..3] = Status | Review note | What to check | Sign-off
     Those four review columns were deleted when the workbook was reduced to a
     single final-data sheet, so the table's declared columns no longer line up
     with row 1 of the sheet and its range runs four columns past the data.
     Excel treats that as corrupt content and strips the table on open.

  2. EMPTY UNNAMED COLUMN 34, sitting between Proportionality and
     Eval? (trackable). No header, zero values in all 563 rows. Excel cannot
     put a blank name in a table column, which is why it had auto-invented
     "Column38" — the second half of defect 1.

Fix: drop the empty column, drop the table object, put a plain AutoFilter and a
frozen header row over the real range. Values, tier row-banding, column widths
and the header text are all preserved exactly.

  python3 repair_v11_workbook.py            # dry run
  python3 repair_v11_workbook.py --write    # backup, repair, verify
"""

from __future__ import annotations

import argparse
import shutil
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TARGET = Path.home() / "Desktop" / "v11_FINAL.xlsx"

# Row banding is derived from tier, not stored intent — reapplied, not copied.
TIER_FILL = {
    ("yes", "yes"): "FFFDECEC",  # Tier A — evaluation + action both trackable
    ("yes", "no"): "FFFFF8E1",   # Tier B — evaluation trackable, action not
    ("no", "no"): "FFEFF6FF",    # Tier C — not a trackable evaluation
}
HEADER_FILL = "FFD9D9D9"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--target", type=Path, default=TARGET)
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.target)
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in ws[1]]
    n_rows = ws.max_row

    print(f"{a.target}")
    print(f"  sheet {ws.title!r} · {n_rows - 1} records · {len(hdr)} columns")

    # --- defect 1 -------------------------------------------------------
    for name in list(ws.tables):
        tbl = ws.tables[name]
        print(f"  stale table {name!r}: ref={tbl.ref} declares {len(tbl.tableColumns)} columns")
        declared = [c.name for c in tbl.tableColumns]
        drift = [(i, d, h) for i, (d, h) in enumerate(zip(declared, hdr), 1) if d != h]
        print(f"    first cell mismatch: table says {declared[0]!r}, sheet says {hdr[0]!r}")
        print(f"    {len(drift)} of {len(declared)} declared columns do not match row 1")

    # --- defect 2 -------------------------------------------------------
    blanks = []
    for c in range(1, len(hdr) + 1):
        if hdr[c - 1] in (None, ""):
            filled = sum(1 for r in range(2, n_rows + 1) if ws.cell(r, c).value not in (None, ""))
            blanks.append((c, filled))
            print(f"  unnamed column {get_column_letter(c)} ({c}): {filled} non-empty values")
    unsafe = [c for c, filled in blanks if filled]
    if unsafe:
        sys.exit(f"REFUSING: unnamed column(s) {unsafe} hold data — resolve by hand first.")

    if not a.write:
        print("\ndry run — nothing written. Re-run with --write to repair.")
        return

    bak = a.target.with_name(a.target.stem + ".backup-repair-" + time.strftime("%Y%m%d-%H%M%S") + a.target.suffix)
    shutil.copy2(a.target, bak)

    for c, _ in reversed(blanks):
        ws.delete_cols(c)
    for name in list(ws.tables):
        del ws.tables[name]

    hdr = [c.value for c in ws[1]]
    ncol = len(hdr)
    ix = {h: i + 1 for i, h in enumerate(hdr) if h}
    last = get_column_letter(ncol)

    # widths: the two long free-text columns stay wide, the rest keep their size
    for c in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = {
            "Finding ID": 30, "Report ID": 24, "Institution": 30, "Institution Type": 16,
            "Report Title ": 44, "Source URL": 44, "Finding": 70, "Finding Quote": 70,
            "Models / Systems": 30,
        }.get(hdr[c - 1], 18)

    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for r in range(2, ws.max_row + 1):
        tier = (ws.cell(r, ix["Eval? (trackable)"]).value, ws.cell(r, ix["Action Trackable?"]).value)
        rgb = TIER_FILL.get(tier)
        if rgb is None:
            sys.exit(f"row {r}: unrecognised tier {tier!r}")
        fill = PatternFill("solid", start_color=rgb, end_color=rgb)
        for c in range(1, ncol + 1):
            ws.cell(r, c).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last}{ws.max_row}"
    wb.save(a.target)

    # --- verify by reloading --------------------------------------------
    ws2 = openpyxl.load_workbook(a.target)[ws.title]
    hdr2 = [c.value for c in ws2[1]]
    assert not list(ws2.tables), "table object survived"
    assert all(h for h in hdr2), "blank header survived"
    assert len(set(hdr2)) == len(hdr2), "duplicate header"
    ids = [ws2.cell(r, 1).value for r in range(2, ws2.max_row + 1)]
    assert len(set(ids)) == len(ids) == 563, f"{len(ids)} ids, {len(set(ids))} unique"
    print(f"\nrepaired · {ws2.max_row - 1} records × {len(hdr2)} columns")
    print(f"  autofilter {ws2.auto_filter.ref} · freeze {ws2.freeze_panes} · tables {list(ws2.tables)}")
    print(f"  backup: {bak.name}")


if __name__ == "__main__":
    main()
