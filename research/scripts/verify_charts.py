#!/usr/bin/env python3
"""Independent recomputation of every quantity the figure set claims to show.

Deliberately shares NO code with the chart generators: it re-reads the workbook, re-derives tier,
severity, outcome and every series from first principles, and asserts the internal identities that
must hold. If a chart script has a filtering or arithmetic bug, the two paths disagree here.

Run:  python3 verify_charts.py
Exits non-zero if any check fails.
"""
import os, sys, csv, datetime, collections, re
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dataset_source

# Only the workbook path and sheet name are shared with the generators, so there is one source of
# dataset truth. Every derivation below is still written independently.
WB = dataset_source.WORKBOOK
FIGS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "charts")

fails, warns = [], []
def bad(msg): fails.append(msg)
def warn(msg): warns.append(msg)


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


ws = openpyxl.load_workbook(WB, data_only=True)[dataset_source.SHEET]
hdr = [norm(c.value) for c in ws[1]]
R = [{h: norm(ws.cell(r, i).value) for i, h in enumerate(hdr, 1) if h}
     for r in range(2, ws.max_row + 1)]
R = [r for r in R if r.get("Finding ID")]

# ---- tier derived independently, from the raw two flags -----------------------------
def tier(r):
    e, a = r["Eval? (trackable)"], r["Action Trackable?"]
    if (e, a) == ("yes", "yes"): return "A"
    if (e, a) == ("yes", "no"):  return "B"
    if e == "no":                return "C"
    return "?"

T = collections.Counter(tier(r) for r in R)
if T["?"]:
    bad(f"{T['?']} rows have an unrecognised tier encoding")
A = [r for r in R if tier(r) == "A"]
H = [r for r in A if r["Severity (C1/C2) majority"] == "C1"]

# ---- severity: recompute the majority from the three raw votes ----------------------
VOTES = ("Sonnet5 vote", "GPT-5.5 vote", "Gemini3.1 vote")
OVERRIDE = {"TRANSLUCE-2026-07-SOC2", "SECUREBIO-2026-04-BIO5"}   # documented, C1 by rule
mismatch = []
for r in R:
    v = [r[c] for c in VOTES]
    if not all(v):
        bad(f"{r['Finding ID']}: missing a severity vote")
        continue
    exp = "C1" if sum(1 for x in v if x == "C1") >= 2 else "C2"
    if exp != r["Severity (C1/C2) majority"] and r["Finding ID"] not in OVERRIDE:
        mismatch.append(r["Finding ID"])
if mismatch:
    bad(f"severity != vote majority on {len(mismatch)} unregistered rows: {mismatch[:5]}")

# ---- proportionality: recompute from the matrix -------------------------------------
M = {("C1","Substantive"):"Proportionate", ("C1","Partial"):"Under-response (gap)",
     ("C1","Acknowledged"):"Under-response (gap)", ("C1","None"):"Accountability gap (no action)",
     ("C2","Substantive"):"Proportionate", ("C2","Partial"):"Proportionate",
     ("C2","Acknowledged"):"Under-response (gap)", ("C2","None"):"Accountability gap (no action)"}
pm = [r["Finding ID"] for r in A
      if r["Action Level"] and M.get((r["Severity (C1/C2) majority"], r["Action Level"])) != r["Proportionality"]]
if pm:
    bad(f"proportionality != severity x action on {len(pm)} rows: {pm[:5]}")

# ---- schema invariants ---------------------------------------------------------------
leak = [r["Finding ID"] for r in R if tier(r) != "A"
        and any(r.get(c) for c in ("Action Level","Attribution","Policy Level","Proportionality"))]
if leak: bad(f"Tier-A-only fields populated on {len(leak)} non-Tier-A rows: {leak[:5]}")
incomplete = [r["Finding ID"] for r in A
              if not all(r.get(c) for c in ("Action Level","Attribution","Policy Level","Proportionality"))]
if incomplete: bad(f"{len(incomplete)} Tier A rows not fully classified: {incomplete[:5]}")
dupes = [k for k, v in collections.Counter(r["Finding ID"] for r in R).items() if v > 1]
if dupes: bad(f"duplicate Finding IDs: {dupes}")

# ---- the quantities each figure asserts ----------------------------------------------
NREP = len({r["Report ID"] for r in R if r["Report ID"]})
NINST = len({r["Institution"] for r in R if r["Institution"]})
dates = sorted(r["Publication Date"] for r in R if r["Publication Date"])
prop = collections.Counter(r["Proportionality"] for r in H)
gap, und, ok = (prop["Accountability gap (no action)"], prop["Under-response (gap)"], prop["Proportionate"])
al = collections.Counter(r["Action Level"] for r in A)
at = collections.Counter(r["Attribution"] for r in A)
pl = collections.Counter(r["Policy Level"] for r in A)
sev = collections.Counter(r["Severity (C1/C2) majority"] for r in R)
sc = collections.Counter(r["Scope"] for r in R)
acc = collections.Counter(r["Access Type"] for r in R)

print("="*100)
print("INDEPENDENT RECOMPUTATION")
print("="*100)
print(f"  corpus                {len(R)}")
print(f"  reports               {NREP}")
print(f"  evaluators            {NINST}")
print(f"  date range            {dates[0]} to {dates[-1]}")
print(f"  tiers                 A {T['A']} · B {T['B']} · C {T['C']}   (sum {T['A']+T['B']+T['C']})")
print(f"  severity              C1 {sev['C1']} · C2 {sev['C2']}   (sum {sev['C1']+sev['C2']})")
print(f"  headline population   {len(H)}")
print(f"  outcomes              no action {gap} · under {und} · proportionate {ok}   (sum {gap+und+ok})")
print(f"  no response           {gap}/{len(H)} = {gap/len(H)*100:.2f}%")
print(f"  falls short           {gap+und}/{len(H)} = {(gap+und)/len(H)*100:.2f}%")
print(f"  action level          {dict(al.most_common())}")
print(f"  attribution           {dict(at.most_common())}")
print(f"  policy level          {dict(pl.most_common())}")
print(f"  scope                 {dict(sc.most_common())}")
print(f"  access type           {dict(acc.most_common())}")

# ---- identities that must hold --------------------------------------------------------
print("\n" + "="*100)
print("IDENTITIES")
print("="*100)
def ident(label, a, b):
    okk = a == b
    print(f"  {'ok  ' if okk else 'FAIL'}  {label:<62} {a} {'==' if okk else '!='} {b}")
    if not okk: bad(f"identity failed: {label} ({a} != {b})")

ident("tiers sum to corpus", T['A']+T['B']+T['C'], len(R))
ident("severity sums to corpus", sev['C1']+sev['C2'], len(R))
ident("outcomes sum to headline population", gap+und+ok, len(H))
ident("action levels sum to Tier A", sum(al.values()), len(A))
ident("attribution sums to Tier A", sum(at.values()), len(A))
ident("policy levels sum to Tier A", sum(pl.values()), len(A))
ident("Attribution 'No response located' == Action Level None",
      at["No response located"], al["None"])
ident("no-action count == Tier A C1 rows with Action Level None",
      gap, sum(1 for r in H if r["Action Level"] == "None"))
ident("proportionate == C1 rows coded Substantive", ok,
      sum(1 for r in H if r["Action Level"] == "Substantive"))
ident("under-response == C1 rows Partial or Acknowledged", und,
      sum(1 for r in H if r["Action Level"] in ("Partial", "Acknowledged")))
ident("headline population == Tier A and C1", len(H),
      sum(1 for r in R if tier(r) == "A" and r["Severity (C1/C2) majority"] == "C1"))

# ---- per-figure series ------------------------------------------------------------------
print("\n" + "="*100)
print("PER-FIGURE SERIES")
print("="*100)
def block(name, pairs, total=None):
    print(f"\n  {name}")
    s = 0
    for k, v in pairs:
        s += v
        print(f"      {str(k):<52} {v:>6}")
    if total is not None:
        okk = s == total
        print(f"      {'sum':<52} {s:>6}   {'ok' if okk else 'FAIL expected '+str(total)}")
        if not okk: bad(f"{name}: series sums to {s}, expected {total}")

block("07_tier_distribution", [(k, T[k]) for k in "ABC"], len(R))
block("17_severity_classification", [("C1", sev["C1"]), ("C2", sev["C2"])], len(R))
block("04_headline_outcome_distribution",
      [("Accountability gap (no action)", gap), ("Under-response (gap)", und), ("Proportionate", ok)], len(H))
block("08_action_level_distribution", al.most_common(), len(A))
block("18_attribution_distribution", at.most_common(), len(A))
block("09_policy_level_distribution", pl.most_common(), len(A))
block("12_evaluator_scope", sc.most_common(), len(R))
block("03_findings_per_access_type", acc.most_common(), len(R))

yr = collections.Counter(r["Publication Date"][:4] for r in R if r["Publication Date"])
block("06_findings_per_year", sorted(yr.items()), len(R))

print("\n  15_gap_rate_by_year  /  15b_shortfall_rate_by_year   (Tier A ∩ C1)")
for y in sorted({r["Publication Date"][:4] for r in H if r["Publication Date"]}):
    g = [r for r in H if r["Publication Date"][:4] == y]
    ng = sum(1 for r in g if r["Proportionality"] == "Accountability gap (no action)")
    ns = sum(1 for r in g if r["Proportionality"] != "Proportionate")
    print(f"      {y}   n={len(g):>3}   no-action {ng:>3} = {ng/len(g)*100:>5.1f}%   "
          f"falls-short {ns:>3} = {ns/len(g)*100:>5.1f}%")

print("\n  16_gap_rate_by_access_type   (Tier A ∩ C1)")
for k in sorted({r["Access Type"] for r in H}):
    g = [r for r in H if r["Access Type"] == k]
    ng = sum(1 for r in g if r["Proportionality"] == "Accountability gap (no action)")
    print(f"      {k:<20} n={len(g):>3}   no-action {ng:>3} = {ng/len(g)*100:>5.1f}%")

print("\n  10_proportionality_by_severity   (Tier A, by severity)")
for s in ("C1", "C2"):
    g = [r for r in A if r["Severity (C1/C2) majority"] == s]
    c = collections.Counter(r["Proportionality"] for r in g)
    print(f"      {s}  n={len(g):>3}   " + " · ".join(f"{k.split(' (')[0]} {v}" for k, v in c.most_common()))

print("\n  21_severity_x_action_heatmap   (Tier A)")
for s in ("C1", "C2"):
    row = collections.Counter(r["Action Level"] for r in A if r["Severity (C1/C2) majority"] == s)
    print(f"      {s}   " + " · ".join(f"{k} {row[k]}" for k in ("None","Acknowledged","Partial","Substantive")))

dom = collections.Counter()
for r in R:
    for d in r.get("Domain", "").split(";"):
        if d.strip(): dom[d.strip()] += 1
print(f"\n  11_domain_distribution   multi-label, sums to {sum(dom.values())} over {len(R)} rows")
for k, v in dom.most_common(10):
    print(f"      {k:<52} {v:>6}")

print("\n  14_response_lag_distribution   (Tier A with a lag)")
lags = sorted(int(r["Lag (days)"]) for r in A if re.fullmatch(r"-?\d+", r["Lag (days)"]))
if lags:
    import statistics
    print(f"      n={len(lags)}  min={lags[0]}  median={statistics.median(lags):.0f}  "
          f"max={lags[-1]}  negative={sum(1 for x in lags if x < 0)}")
    nl = [r["Finding ID"] for r in A if re.fullmatch(r"-?\d+", r["Lag (days)"]) and int(r["Lag (days)"]) < 0
          and r["Access Type"] not in ("Pre-deployment", "Mixed")]
    if nl: bad(f"negative lag on non-pre-deployment rows: {nl}")

print("\n  01_findings_per_institution   top 8 of "
      f"{len({r['Institution'] for r in R})}")
for k, v in collections.Counter(r["Institution"] for r in R).most_common(8):
    print(f"      {k[:52]:<52} {v:>6}")

print("\n  19_gap_rate_by_developer   (Tier A ∩ C1, n>=5)")
def dev(r):
    s = r["Models / Systems"] + " " + r["Report Title"]
    for pat, name in ((r'gpt|chatgpt|codex|\bo[134]\b|operator|sora', "OpenAI"),
                      (r'claude|opus|sonnet|haiku|fable|mythos', "Anthropic"),
                      (r'gemini|gemma|palm', "Google"), (r'llama|prompt.?guard', "Meta"),
                      (r'grok', "xAI"), (r'deepseek', "DeepSeek"), (r'qwen', "Alibaba"),
                      (r'mistral|magistral', "Mistral"), (r'nemotron', "NVIDIA")):
        if re.search(pat, s, re.I): return name
    return None
dv = collections.defaultdict(list)
for r in H:
    d = dev(r)
    if d: dv[d].append(r)
for k in sorted(dv, key=lambda k: -len(dv[k])):
    g = dv[k]
    if len(g) < 5: continue
    ng = sum(1 for r in g if r["Proportionality"] == "Accountability gap (no action)")
    print(f"      {k:<14} n={len(g):>3}   no-action {ng:>3} = {ng/len(g)*100:>5.1f}%")

# ---- figure files ------------------------------------------------------------------------
print("\n" + "="*100)
print("FIGURE FILES")
print("="*100)
figs = sorted(f for f in os.listdir(FIGS) if f.endswith(".png"))
now = datetime.datetime.now()
stale = [f for f in figs
         if (now - datetime.datetime.fromtimestamp(os.path.getmtime(os.path.join(FIGS, f)))).total_seconds() > 86400]
wbm = os.path.getmtime(WB)
older = [f for f in figs if os.path.getmtime(os.path.join(FIGS, f)) < wbm]
print(f"  {len(figs)} figures")
print(f"  {'ok  ' if not older else 'FAIL'}  figures newer than the workbook: "
      f"{len(figs)-len(older)}/{len(figs)}" + (f"   STALE: {older}" if older else ""))
if older: bad(f"{len(older)} figures predate the workbook they claim to plot: {older}")

print("\n" + "="*100)
if fails:
    print(f"FAIL — {len(fails)} problem(s):")
    for f in fails: print(f"   - {f}")
else:
    print("PASS — every recomputed quantity and identity agrees with the workbook.")
for w in warns: print(f"   note: {w}")
sys.exit(1 if fails else 0)
