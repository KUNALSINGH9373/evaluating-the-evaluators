#!/usr/bin/env python3
"""26_action_level_scale.png — how company response strength is measured.

Four rows, strongest first. Each level carries its meaning inside its own block, then two
columns: the finding, and what the company did about it.

Deliberately minimal — no counts, no footnote, no finding IDs, no derived-outcome line. This
chart answers "how is strength measured"; distribution and derivation have their own figures.
Each example is still checked against the workbook at build time, so it cannot drift from the
row it illustrates even though the row is no longer named on the figure.
"""
import os, datetime
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch
import openpyxl
import dataset_source
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palette import *   # single source of colour truth


OUT = CHARTS_OUT
plt.rcParams.update({"figure.dpi": 200, "savefig.dpi": 200, "font.family": "DejaVu Sans",
                     "figure.facecolor": "white", "axes.facecolor": "white",
                     "savefig.facecolor": "white"})
INK, SOFT, FAINT = "#111111", "#4A4A4A", "#8A959F"
FIG_W = 13.6   # narrow on purpose: a wide canvas left the text columns short and airy


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


ws = dataset_source.sheet()
hdr = [norm(c.value) for c in ws[1]]
ROWS = {r["Finding ID"]: r for r in
        ({h: norm(ws.cell(i, k).value) for k, h in enumerate(hdr, 1) if h}
         for i in range(2, ws.max_row + 1)) if r.get("Finding ID")}

SHORT_PROP = {"Proportionate": "Proportionate", "Under-response (gap)": "Under-response",
              "Accountability gap (no action)": "Accountability gap"}

LEVELS = [
    ("Substantive", "A specific mitigation, safeguard or deployment decision",
     "JOINT-2025-09-CYB1",
     "CAISI found two novel security vulnerabilities in ChatGPT Agent and chained them "
     "into a working exploit.",
     "OpenAI fixed both within one business day of the report."),
    ("Partial", "Interim or incomplete — or claimed but unverifiable",
     "UKAISI-2026-04-JAI1",
     "UK AISI found a universal jailbreak in GPT-5.5 within six hours.",
     "OpenAI said it had patched it, but AISI could not verify the fix before public deployment."),
    ("Acknowledged", "The problem is referenced, no action specified",
     "ANTHROPIC-2025-09-SELF-ALI1",
     "UK AISI found Claude Sonnet 4.5 verbalises unprompted evaluation awareness in 16.6% of "
     "outputs, against 1.6% for Opus 4.1.",
     "Anthropic reproduced the measurement in the system card, but recorded no safeguard change."),
    ("None", "No public company response located",
     "DREADNODE-2026-07-ALI1",
     "Dreadnode found Anthropic's Opus 4.8 and Sonnet 5 were the heaviest cheaters of the 22 "
     "models tested (65.2% and 56.5%).",
     "Five-source battery exhausted and dated; no company document located."),
]
for name, _, fid, _, _ in LEVELS:
    got = ROWS[fid]["Action Level"]
    assert got == name, f"{fid} is coded {got!r}, not {name!r} — example and level disagree"


def render(fig_h):
    fig = plt.figure(figsize=(FIG_W, fig_h))
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, 100); ax.set_ylim(100, 0); ax.axis("off")
    fig.canvas.draw()
    rend = fig.canvas.get_renderer()
    u = lambda inches: inches * 100.0 / fig_h
    em = lambda fs: u(fs / 72.0)

    def tw(s, fs, weight="normal"):
        t = ax.text(0, 0, s, fontsize=fs, fontweight=weight, alpha=0)
        bb = t.get_window_extent(renderer=rend); t.remove()
        inv = ax.transData.inverted()
        return abs(inv.transform((bb.width, 0))[0] - inv.transform((0, 0))[0])

    def wrap(s, fs, weight, maxw):
        out, cur = [], ""
        for word in s.split(" "):
            trial = f"{cur} {word}".strip()
            if cur and tw(trial, fs, weight) > maxw:
                out.append(cur); cur = word
            else:
                cur = trial
        if cur:
            out.append(cur)
        return out

    def para(x, y, s, fs, weight="normal", colour=INK, maxw=40, lead=1.34):
        for line in wrap(s, fs, weight, maxw):
            ax.text(x, y, line, fontsize=fs, fontweight=weight, color=colour,
                    va="baseline", zorder=3)
            y += em(fs) * lead
        return y

    ax.text(2.0, em(34) * 0.84, "How company response strength is measured",
            fontsize=34, fontweight="bold", color=INK, va="baseline")
    ax.text(2.0, em(34) * 0.84 + em(19) * 1.70,
            "Action Level — the strength of the located company response, with a Tier A example "
            "of each.", fontsize=19, color=SOFT, va="baseline")

    BX, BW = 2.0, 25.0        # level block: name and meaning
    FX, FW = 29.5, 33.5       # the finding
    CX, CW = 65.5, 33.0       # the company response
    PADB = u(0.17)
    GAP = u(0.11)

    head = em(34) * 0.84 + em(19) * 1.70 + u(0.34)
    ax.text(FX, head, "THE FINDING", fontsize=13, fontweight="bold", color=FAINT,
            va="baseline", zorder=3)
    ax.text(CX, head, "WHAT THE COMPANY DID", fontsize=13, fontweight="bold", color=FAINT,
            va="baseline", zorder=3)
    top = head + u(0.16)

    y = top
    for name, desc, fid, finding, company in LEVELS:
        col = ACTION[name]
        # measure all three columns, then give the row the height the tallest one needs
        nb = len(wrap(desc, 15, "normal", BW - 3.4))
        hb = em(26) * 1.08 + em(15) * 1.32 * nb + 2 * PADB
        ht = max(len(wrap(finding, 16, "normal", FW)),
                 len(wrap(company, 16, "normal", CW))) * em(16) * 1.32 + 2 * PADB
        h = max(hb, ht, u(0.85))

        ax.add_patch(FancyBboxPatch((BX, y), BW, h,
                                    boxstyle="round,pad=0,rounding_size=0.55",
                                    facecolor=col, edgecolor="none", zorder=1, clip_on=False))
        block_h = em(26) * 1.08 + em(15) * 1.32 * nb
        by = y + (h - block_h) / 2 + em(26) * 0.78
        ax.text(BX + 1.5, by, name, fontsize=26, fontweight="bold", color="white",
                va="baseline", zorder=3)
        para(BX + 1.5, by + em(26) * 1.00, desc, 15, "normal", "#FFFFFF", BW - 3.0, 1.32)

        # centre each column's own block of lines: the row is as tall as its tallest element,
        # and dumping the shorter columns at the top is what read as wasted space
        for x, s, wdt in ((FX, finding, FW), (CX, company, CW)):
            n = len(wrap(s, 16, "normal", wdt))
            para(x, y + (h - n * em(16) * 1.32) / 2 + em(16) * 0.80, s, 16, "normal", INK,
                 wdt, 1.32)
        y += h + GAP

    ax.annotate("", xy=(1.0, y - GAP - u(0.35)), xytext=(1.0, top + u(0.35)),
                arrowprops=dict(arrowstyle="-|>", color="#C3CAD2", lw=3.0,
                                shrinkA=0, shrinkB=0, mutation_scale=24))
    ax.text(0.15, (top + y) / 2, "W E A K E R", rotation=90, ha="center", va="center",
            fontsize=12, color=FAINT, fontweight="bold")
    return fig, (y - GAP) * fig_h / 100.0


probe, need = render(9.0)
plt.close(probe)
fig, _ = render(need + 0.14)
p = os.path.join(OUT, "26_action_level_scale.png")
fig.savefig(p, bbox_inches="tight", pad_inches=0.20)
plt.close(fig)
print(f"wrote {p}")
for name, _, fid, _, _ in LEVELS:
    r = ROWS[fid]
    print(f"  {name:<13} {fid:<28} {r['Severity (C1/C2) majority']} -> {r['Proportionality']}")
    # IDs are checked here but no longer printed on the figure
