#!/usr/bin/env python3
"""25_what_is_a_finding.png — the unit of analysis, on one card.

Four panels: the definition, the venues enumerated, the window, the corpus totals. Every number
is read from the workbook, so the card cannot drift from the dataset.

Deliberately spare — composition breakdowns (finding type, domain, access) have their own figures,
and repeating them here buried the one number this card exists to deliver.

Text is laid out by a small flow engine rather than by hand-placed coordinates: line width and
line height are measured off the renderer, so copy can be edited without lines colliding or
running past a panel edge. Hand-placed y-offsets were tried first and overflowed on every panel.
"""
import os, datetime
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch
import openpyxl
import dataset_source
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palette import *   # single source of colour truth


OUT = CHARTS_OUT
plt.rcParams.update({"figure.dpi": 200, "savefig.dpi": 200, "font.family": "DejaVu Sans",
                     "figure.facecolor": "white", "axes.facecolor": "white",
                     "savefig.facecolor": "white"})
INK, SOFT, FAINT, RULE = "#111111", "#4A4A4A", "#6E7A85", "#D3DAE1"
PANEL = "#F3F6F9"
ACCENT = INSTTYPE["Government"]
FIG_W = 10.6   # portrait: a narrow column, height solved for by the two-pass render


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


ws = dataset_source.sheet()
hdr = [norm(c.value) for c in ws[1]]
R = [{h: norm(ws.cell(r, i).value) for i, h in enumerate(hdr, 1) if h}
     for r in range(2, ws.max_row + 1)]
R = [r for r in R if r.get("Finding ID")]

N_FIND = len(R)
N_REP = len({r["Report ID"] for r in R if r["Report ID"]})
N_ORG = len({r["Institution"] for r in R})
PER_REP = N_FIND / N_REP
YEARS = sorted({r["Publication Date"][:4] for r in R if r["Publication Date"]})
BY_YEAR = [(y,
            len({r["Report ID"] for r in R if r["Publication Date"][:4] == y and r["Report ID"]}),
            sum(1 for r in R if r["Publication Date"][:4] == y))
           for y in YEARS]
FIRST = min(r["Publication Date"] for r in R if r["Publication Date"])
LAST = max(r["Publication Date"] for r in R if r["Publication Date"])
MONTH = ["January", "February", "March", "April", "May", "June", "July", "August",
         "September", "October", "November", "December"]
span = f"{MONTH[int(FIRST[5:7]) - 1]} {FIRST[:4]} – {MONTH[int(LAST[5:7]) - 1]} {LAST[:4]}"
cutoff = f"{int(LAST[8:10])} {MONTH[int(LAST[5:7]) - 1]} {LAST[:4]}"


def render(fig_h):
    """Draw the whole card at a given figure height; return (fig, content bottom in inches).

    Two passes are used. Vertical spacing is specified in INCHES and converted to grid units by
    u(), so the layout's physical height is identical whatever canvas it is drawn on — which
    means one measuring pass gives the exact height the second pass should use. Specifying the
    spacing in grid units instead made the layout grow as the canvas grew, and it never converged.
    """
    fig = plt.figure(figsize=(FIG_W, fig_h))
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, 100); ax.set_ylim(100, 0); ax.axis("off")
    fig.canvas.draw()
    rend = fig.canvas.get_renderer()

    def u(inches):
        return inches * 100.0 / fig_h

    def em(fs):
        return u(fs / 72.0)

    def tw(s, fs, weight="normal"):
        t = ax.text(0, 0, s, fontsize=fs, fontweight=weight, alpha=0)
        bb = t.get_window_extent(renderer=rend)
        t.remove()
        inv = ax.transData.inverted()
        return abs(inv.transform((bb.width, 0))[0] - inv.transform((0, 0))[0])

    def wrap(s, fs, weight, maxw):
        out, cur = [], ""
        for word in s.split(" "):
            trial = f"{cur} {word}".strip()
            if cur and tw(trial, fs, weight) > maxw:
                out.append(cur); cur = word
            else:
                cur = trial
        if cur:
            out.append(cur)
        return out

    def para(x, y, s, fs, weight="normal", colour=INK, maxw=40, lead=1.32):
        for line in wrap(s, fs, weight, maxw):
            ax.text(x, y, line, fontsize=fs, fontweight=weight, color=colour,
                    va="baseline", zorder=3)
            y += em(fs) * lead
        return y

    # Portrait, single column. Four panels stacked, each sized to its own content — nothing is
    # padded out to match a neighbour, which is what a two-column grid forces on the short panels.
    X0, X1, PADX = 2.0, 98.0, 2.6
    PADB, GUTY = u(0.26), u(0.16)
    TITLES = {1: "Definition", 2: "The corpus", 3: "Venues", 4: "Timeline"}
    W = X1 - X0 - 2 * PADX

    def panel(i, y0, y1):
        ax.add_patch(FancyBboxPatch((X0, y0), X1 - X0, y1 - y0,
                                    boxstyle="round,pad=0,rounding_size=0.6",
                                    facecolor=tint(ACCENT, 0.90) if i == 2 else PANEL,
                                    edgecolor="none", zorder=0, clip_on=False))
        ax.add_patch(FancyBboxPatch((X0, y0), 0.55, y1 - y0,
                                    boxstyle="round,pad=0,rounding_size=0.22",
                                    facecolor=ACCENT, edgecolor="none", zorder=1, clip_on=False))
        ax.text(X0 + PADX, y0 + u(0.30) + em(26) * 0.72, TITLES[i], fontsize=26,
                fontweight="bold", color=ACCENT, va="baseline", zorder=3)

    x = X0 + PADX
    ax.text(X0, em(42) * 0.80, "What counts as a finding", fontsize=42, fontweight="bold",
            color=INK, va="baseline", zorder=3)
    TOP = em(42) * 0.80 + u(0.30)
    HEAD = u(0.30) + em(26) * 0.72 + em(26) * 1.30   # panel title row

    # ---- 1 · definition -----------------------------------------------------------------
    y = TOP + HEAD
    y = para(x, y, "A discrete, evidence-backed claim asserted by a government AI institute or "
             "an independent evaluation organisation in a public report, codeable independently "
             "of the report's other claims.", 23, "bold", INK, W, 1.30) + u(0.26)
    y = para(x, y, "Qualifies:  empirical model findings, including reassuring nulls  ·  "
             "methodology and tooling  ·  governance and process",
             19, "normal", INK, W, 1.30) + u(0.13)
    y = para(x, y, "Not findings:  announcements  ·  partnership news  ·  opinion pieces  "
             "·  plans", 19, "normal", FAINT, W, 1.30)
    b = y + PADB
    panel(1, TOP, b)

    # ---- 2 · corpus ---------------------------------------------------------------------
    # Big number on the left, the two smaller counts stacked beside it — a full-width row per
    # statistic would have left most of the panel empty.
    top2 = b + GUTY
    y = top2 + HEAD
    # The only figure on this card NOT derived from the workbook: the screening ledger lives in
    # sweep_state/master_ledger.csv, outside this repo, so it cannot be recomputed here.
    # Rulebook §3 records 6,579 rows, 86.6% resolved. Stated as "6,500+" rather than exactly,
    # because the ledger's retirement arithmetic does not currently reconcile to the item.
    ax.text(x, y, "6,500+ publications screened", fontsize=24, fontweight="bold",
            color=ACCENT, va="baseline", zorder=3)
    y += em(84) * 0.74 + em(24) * 0.60
    ax.text(x, y, f"{N_FIND:,}", fontsize=84, fontweight="bold", color=INK,
            va="baseline", zorder=3)
    ax.text(x, y + em(22) * 1.55, "findings", fontsize=22, color=SOFT, va="baseline", zorder=3)
    sx = x + tw(f"{N_FIND:,}", 84, "bold") + 6.0
    ax.plot([sx - 3.0, sx - 3.0], [y - em(84) * 0.70, y + em(22) * 1.75],
            color=tint(ACCENT, 0.5), lw=1.8, zorder=2)
    sy = y - em(84) * 0.34
    for big, small in ((f"{N_REP}", "reports"), (f"{N_ORG}", "evaluator organisations")):
        ax.text(sx, sy, big, fontsize=32, fontweight="bold", color=ACCENT,
                va="baseline", zorder=3)
        ax.text(sx + tw(big, 32, "bold") + 2.2, sy, small, fontsize=21, color=INK,
                va="baseline", zorder=3)
        sy += em(32) * 1.70
    b = max(y + em(22) * 1.55, sy - em(32) * 0.70) + PADB
    panel(2, top2, b)

    # ---- 3 · venues ---------------------------------------------------------------------
    top3 = b + GUTY
    y = top3 + HEAD
    y = para(x, y, "Publication index   ·   arXiv   ·   Articles   ·   Reports   ·   "
             "System cards", 21, "normal", INK, W, 1.35)
    b = y - em(21) * 0.34 + PADB
    panel(3, top3, b)

    # ---- 4 · timeline -------------------------------------------------------------------
    top4 = b + GUTY
    y = top4 + HEAD + em(52) * 0.74
    ax.text(x, y, f"{FIRST[:4]} – {LAST[:4]}", fontsize=52, fontweight="bold", color=INK,
            va="baseline", zorder=3)
    ax.text(x + tw(f"{FIRST[:4]} – {LAST[:4]}", 52, "bold") + 3.4, y,
            f"Corpus frozen {cutoff}", fontsize=21, color=SOFT, va="baseline", zorder=3)
    row2 = y + PADB
    panel(4, top4, row2)


    return fig, row2 * fig_h / 100.0


TRIAL = 13.0
probe, need = render(TRIAL)
plt.close(probe)
fig, _ = render(need + 0.22)

p = os.path.join(OUT, "25_what_is_a_finding.png")
fig.savefig(p, bbox_inches="tight", pad_inches=0.3)
plt.close(fig)
print(f"wrote {p}   {N_FIND:,} findings · {N_REP} reports · {N_ORG} orgs · {PER_REP:.2f}/report")
print("  " + " · ".join(f"{y}: {a} reports / {b} findings" for y, a, b in BY_YEAR))
