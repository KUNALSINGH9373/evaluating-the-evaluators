#!/usr/bin/env python3
"""Sort the sheet to the rulebook schema.

v11_RULEBOOK.md, Schema:
    "rows sorted Tier A, Tier B, Tier C and then newest-to-oldest within each tier."

Tier is DERIVED from Eval? (trackable) + Action Trackable?, never read from a label.
Within a tier, rows go newest-to-oldest; Report ID then Finding ID break ties so a report's
findings stay adjacent.

I previously sorted this sheet by date alone, which ignored the tier blocking entirely and left
101 tier runs. This restores the documented order.

Note: the original V12 sheet was itself only partly compliant (5 tier runs, and Tier A not in
date order), so applying the rule moves some V12 rows relative to each other. That is what the
rule requires. Cell contents are asserted unchanged before saving.
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dataset_source as ds
import openpyxl

def natkey(s):
    return [int(t) if t.isdigit() else t for t in re.split(r'(\d+)', s or "")]

RANK = {"A": 0, "B": 1, "C": 2}

wb = openpyxl.load_workbook(ds.WORKBOOK)
ws = wb[ds.SHEET]
hdr = [ds.norm(c.value) for c in ws[1]]
ncol = len(hdr)
i_id, i_rep, i_date = hdr.index("Finding ID"), hdr.index("Report ID"), hdr.index("Publication Date")
i_ev, i_at = hdr.index("Eval? (trackable)"), hdr.index("Action Trackable?")

body = [[ws.cell(r, c).value for c in range(1, ncol + 1)]
        for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value]
snap = {ds.norm(r[i_id]): tuple(ds.norm(v) for v in r) for r in body}
assert len(snap) == len(body), "duplicate Finding ID — refusing to sort"

def tier(row):
    e, a = ds.norm(row[i_ev]), ds.norm(row[i_at])
    return "A" if e == "yes" and a == "yes" else "B" if e == "yes" else "C"

body.sort(key=lambda row: (
    RANK[tier(row)],
    tuple(-ord(ch) for ch in ds.norm(row[i_date])[:10].ljust(10)),   # newest first
    natkey(ds.norm(row[i_rep])),
    natkey(ds.norm(row[i_id])),
))

after = {ds.norm(r[i_id]): tuple(ds.norm(v) for v in r) for r in body}
assert after == snap, "cell contents changed — aborting"

for i, row in enumerate(body):
    for c, v in enumerate(row, start=1):
        ws.cell(i + 2, c).value = v
wb.save(ds.WORKBOOK)

import itertools
seq = [tier(r) for r in body]
runs = [k for k, _ in itertools.groupby(seq)]
print(f"wrote {len(body)} rows · tier runs: {len(runs)} ({''.join(runs)})")
for t in "ABC":
    d = [ds.norm(r[i_date])[:10] for r in body if tier(r) == t]
    print(f"   Tier {t}: n={len(d):>4}  newest-to-oldest: {d == sorted(d, reverse=True)}   {d[0]} .. {d[-1]}")
