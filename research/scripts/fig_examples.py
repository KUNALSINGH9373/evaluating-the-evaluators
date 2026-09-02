#!/usr/bin/env python3
"""27_three_outcomes.png — three 2026 findings as a plain table.

Five columns, in the order specified: finding, company response, policy response, company action
level, policy action level. No cards, no badges, no severity or derived-outcome column.

All three rows are Tier A and unanimously C1 — severity is held constant off-figure so that the
company response is the only thing varying down the table. The C1 and 2026 conditions are still
asserted against the workbook at build time even though neither is printed.
"""
import os, datetime
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
import dataset_source
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palette import *   # single source of colour truth


OUT = CHARTS_OUT
plt.rcParams.update({"figure.dpi": 200, "savefig.dpi": 200, "font.family": "DejaVu Sans",
                     "figure.facecolor": "white", "axes.facecolor": "white",
                     "savefig.facecolor": "white"})
INK, SOFT, FAINT, RULE = "#111111", "#4A4A4A", "#8A959F", "#C9D3DC"
FIG_W = 17.0


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


ws = dataset_source.sheet()
hdr = [norm(c.value) for c in ws[1]]
ROWS = {r["Finding ID"]: r for r in
        ({h: norm(ws.cell(i, k).value) for k, h in enumerate(hdr, 1) if h}
         for i in range(2, ws.max_row + 1)) if r.get("Finding ID")}

POL_SHORT = {"Binding policy action": "Binding",
             "Non-binding policy-related uptake": "Non-binding",
             "No policy uptake identified": "No uptake identified"}

CASES = [
    ("METR-2026-06-ALI1",
     "METR found GPT-5.6 Sol had the highest observed rate of cheating and reward-hacking of any "
     "model it had evaluated — exploiting bugs in the evaluation environment or adopting "
     "disallowed strategies.",
     "The system card names METR and describes the finding. After the same behaviour recurred in "
     "the wild, OpenAI paused internal deployment and set out concrete mitigations before "
     "restoring limited access after retesting.",
     "UK AISI published a blog post discussing the cheating finding and its implications for "
     "evaluation methodology, recommending models be trained not to cheat."),
    ("UKAISI-2026-04-JAI1",
     "UK AISI found a universal jailbreak in GPT-5.5 within six hours.",
     "OpenAI conducted pre-deployment testing with UK AISI, recorded in the GPT-5.5 system card. "
     "No verified patch of the jailbreak is documented in the cited source.",
     # was "None located." — contradicted the row's own Policy Level of Non-binding uptake once
     # Channel B was re-run. The prose here is hand-written while the levels are read live from the
     # sheet, so the two drift apart silently. Kept in sync with UKAISI-2026-04-JAI1 Policy Response.
     "Three official UK government sources cite the evaluation, all postdating it: a DSIT "
     "press release attributing its cyber-resilience push to \u201crecent research by the AI "
     "Security Institute\u201d, a joint DSIT/NCSC case study, and a ministerial statement."),
    ("UKAISI-2026-04-CYB1",
     "Claude Mythos Preview reached a 73% pass rate on expert-level CTFs and became the first "
     "model to solve a challenge on AISI's hardest “The Last Ones” range.",
     "No publicly documented response located.",
     "A UK government open letter to business leaders from the DSIT Secretary of State and the "
     "Security Minister cites AISI's evaluation, and launched a Cyber Resilience Pledge."),
]
for fid, *_ in CASES:
    assert ROWS[fid]["Severity (C1/C2) majority"] == "C1", f"{fid} is no longer C1"
    assert ROWS[fid]["Publication Date"][:4] == "2026", f"{fid} is not a 2026 finding"

COLS = [("Finding", 24.5), ("Company response", 25.5), ("Policy response", 24.5),
        ("Company action level", 11.5), ("Policy action level", 11.5)]
PADX = 1.1


def render(fig_h):
    fig = plt.figure(figsize=(FIG_W, fig_h))
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, 100); ax.set_ylim(100, 0); ax.axis("off")
    fig.canvas.draw()
    rend = fig.canvas.get_renderer()
    u = lambda inches: inches * 100.0 / fig_h
    em = lambda fs: u(fs / 72.0)

    def tw(s, fs, weight="normal"):
        t = ax.text(0, 0, s, fontsize=fs, fontweight=weight, alpha=0)
        bb = t.get_window_extent(renderer=rend); t.remove()
        inv = ax.transData.inverted()
        return abs(inv.transform((bb.width, 0))[0] - inv.transform((0, 0))[0])

    def wrap(s, fs, weight, maxw):
        out, cur = [], ""
        for word in s.split(" "):
            trial = f"{cur} {word}".strip()
            if cur and tw(trial, fs, weight) > maxw:
                out.append(cur); cur = word
            else:
                cur = trial
        if cur:
            out.append(cur)
        return out

    def para(x, y, s, fs, weight, colour, maxw):
        for line in wrap(s, fs, weight, maxw):
            ax.text(x, y, line, fontsize=fs, fontweight=weight, color=colour,
                    va="baseline", zorder=3)
            y += em(fs) * 1.32
        return y

    X = [1.5]
    for _, w in COLS:
        X.append(X[-1] + w)
    FS, HS, PADY = 15, 12.5, u(0.15)

    y = u(0.10)
    nh = max(len(wrap(c, HS, "bold", w - 2 * PADX)) for c, w in COLS)
    hh = nh * em(HS) * 1.28 + 2 * PADY
    ax.add_patch(plt.Rectangle((X[0], y), X[-1] - X[0], hh, facecolor="#EFF3F7",
                               edgecolor="none", zorder=0))
    for i, (c, w) in enumerate(COLS):
        para(X[i] + PADX, y + PADY + em(HS) * 0.80, c, HS, "bold", "#3A4650", w - 2 * PADX)
    y += hh
    ax.plot([X[0], X[-1]], [y, y], color="#9AA7B2", lw=1.6, zorder=2)

    for fid, finding, company, policy in CASES:
        r = ROWS[fid]
        cells = [finding, company, policy, r["Action Level"], POL_SHORT[r["Policy Level"]]]
        n = max(len(wrap(s, FS, "normal", COLS[i][1] - 2 * PADX)) for i, s in enumerate(cells))
        h = n * em(FS) * 1.32 + 2 * PADY
        for i, s in enumerate(cells):
            # the action-level word keeps its palette colour; everything else is plain ink
            col = shade(ACTION[s], 0.18) if i == 3 else INK
            para(X[i] + PADX, y + PADY + em(FS) * 0.80, s, FS,
                 "bold" if i == 3 else "normal", col, COLS[i][1] - 2 * PADX)
        y += h
        ax.plot([X[0], X[-1]], [y, y], color=RULE, lw=1.0, zorder=2)

    for xi in X:
        ax.plot([xi, xi], [u(0.10), y], color=RULE, lw=1.0, zorder=2)
    ax.plot([X[0], X[-1]], [u(0.10), u(0.10)], color="#9AA7B2", lw=1.6, zorder=2)
    return fig, y * fig_h / 100.0


probe, need = render(8.0)
plt.close(probe)
fig, _ = render(need + 0.10)
p = os.path.join(OUT, "27_three_outcomes.png")
fig.savefig(p, bbox_inches="tight", pad_inches=0.18)
plt.close(fig)
print(f"wrote {p}")
for fid, *_ in CASES:
    r = ROWS[fid]
    print(f"  {fid:<22} {r['Publication Date']}  {r['Severity (C1/C2) majority']}  "
          f"{r['Action Level']:<12} {r['Policy Level']}")
