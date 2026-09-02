#!/usr/bin/env python3
"""12_evaluator_scope.png — evaluator scope, stacked by outcome.

Replaces a plain two-bar count with the question that matters: does the accountability gap differ
between government safety institutes and third-party evaluators? Group sizes are very unequal
(54 vs 93 in the headline population), so the bars are 100% stacked with counts printed inside and
n on the axis — a raw-count stack would make the comparison unreadable.

Reads the AISIEVAL_V12 sheet via dataset_source; writes ../charts/12_evaluator_scope.png.
"""
import os, datetime, collections
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import PercentFormatter
import openpyxl
import dataset_source
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palette import *   # single source of colour truth


ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
plt.rcParams.update({"figure.dpi": 200, "savefig.dpi": 200, "font.family": "DejaVu Sans",
                     "figure.facecolor": "white", "axes.facecolor": "white",
                     "savefig.facecolor": "white"})
# colours come from palette.py


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


ws = dataset_source.sheet()
hdr = [norm(c.value) for c in ws[1]]
ix = {h: i + 1 for i, h in enumerate(hdr) if h}
R = [{h: norm(ws.cell(r, ix[h]).value) for h in ix} for r in range(2, ws.max_row + 1)
     if ws.cell(r, 1).value]
tier = lambda r: ("A" if r["Eval? (trackable)"] == "yes" and r["Action Trackable?"] == "yes"
                  else "B" if r["Eval? (trackable)"] == "yes" else "C")
A = [r for r in R if tier(r) == "A"]
H = [r for r in A if r["Severity (C1/C2) majority"] == "C1"]

SCOPES = [("government-AISI", "Government\nsafety institute"),
          ("third-party-evaluator", "Third-party\nevaluator")]
OUT = [("Accountability gap (no action)", "No documented response", RED),
       ("Under-response (gap)", "Under-response", AMBER),
       ("Proportionate", "Proportionate", GREEN)]

fig, ax = plt.subplots(figsize=(16, 9.2))
y = [1, 0]
labels = []
for yi, (key, disp) in zip(y, SCOPES):
    g = [r for r in H if r["Scope"] == key]
    c = collections.Counter(r["Proportionality"] for r in g)
    labels.append(f"{disp}\n(n={len(g)})")
    left = 0.0
    for okey, olab, col in OUT:
        frac = c[okey] / len(g)
        ax.barh(yi, frac, left=left, color=col, height=0.52, zorder=3,
                edgecolor="white", linewidth=2.5)
        if frac > 0.055:
            ax.text(left + frac / 2, yi, f"{c[okey]}\n{frac:.0%}", ha="center", va="center",
                    fontsize=24, color="white", fontweight="bold", linespacing=1.25)
        left += frac
    # the headline figure for each group, just past the bar
    short = (c["Accountability gap (no action)"] + c["Under-response (gap)"]) / len(g)
    ax.text(1.015, yi, f"{short:.0%}\nfall short", va="center", ha="left",
            fontsize=17, color=RED, fontweight="bold", linespacing=1.3)

ax.set_yticks(y)
ax.set_yticklabels(labels, fontsize=25)
ax.set_xlim(0, 1.145)
ax.set_ylim(-0.55, 1.62)
ax.xaxis.set_major_formatter(PercentFormatter(1.0))
ax.tick_params(axis="x", labelsize=21)
ax.grid(axis="y", visible=False)
ax.grid(axis="x", color=GRID, zorder=0)
for s in ("top", "right", "left"):
    ax.spines[s].set_visible(False)
ax.set_xlabel("Share of the evaluator's significant-risk findings", fontsize=21, labelpad=14)
ax.set_title("Who reports it does not change what happens next",
             fontsize=33, color="#111111", pad=48, loc="left")
ax.text(0, 1.055, "Outcome mix by evaluator type — significant-risk findings that name a company "
        f"(n={len(H)})", transform=ax.transAxes, fontsize=17, color=MUTED, va="bottom")

hand = [plt.Rectangle((0, 0), 1, 1, facecolor=col) for _, _, col in OUT]
ax.legend(hand, [lab for _, lab, _ in OUT], loc="upper center",
          bbox_to_anchor=(0.5, -0.145), ncol=3, frameon=False, fontsize=19)

gov = [r for r in H if r["Scope"] == "government-AISI"]
tp = [r for r in H if r["Scope"] == "third-party-evaluator"]
gs = sum(1 for r in gov if r["Proportionality"] != "Proportionate")
ts = sum(1 for r in tp if r["Proportionality"] != "Proportionate")
fig.text(0.012, 0.012,
         f"Government institutes {gs}/{len(gov)} = {gs/len(gov):.1%} fall short · third-party evaluators "
         f"{ts}/{len(tp)} = {ts/len(tp):.1%}. The headline does not rest on pooling the two: it holds within "
         "each.\nBars are 100% stacked because the groups differ in size; counts are printed inside. "
         f"Corpus-wide the split is {sum(1 for r in R if r['Scope']=='government-AISI')} government "
         f"and {sum(1 for r in R if r['Scope']=='third-party-evaluator')} third-party findings.",
         fontsize=13.5, color=INK_2, linespacing=1.65)
fig.subplots_adjust(left=0.185, right=0.895, top=0.80, bottom=0.30)
p = os.path.join(CHARTS_OUT, "12_evaluator_scope.png")
fig.savefig(p, bbox_inches="tight", pad_inches=0.35)
plt.close(fig)
print(f"wrote {p}")
for key, disp in SCOPES:
    g = [r for r in H if r["Scope"] == key]
    c = collections.Counter(r["Proportionality"] for r in g)
    tot = sum(c[k] for k, _, _ in OUT)
    assert tot == len(g), (tot, len(g))
    print(f"  {key:<24} n={len(g):>3}  " +
          " · ".join(f"{lab} {c[k]} ({c[k]/len(g)*100:.1f}%)" for k, lab, _ in OUT))
