#!/usr/bin/env python3
"""The single source of dataset truth.

AISIEVAL_V13 is the current dataset. Every row, column and cell in it is
authoritative. It is READ IN PLACE and never copied, moved, rewritten or overwritten — every
script imports from here so there is exactly one path and one sheet name in the whole project.

Do not add a second reader. Do not write to this workbook.
"""
import os
import openpyxl
import datetime

# Resolved relative to this file, so the project works from any checkout location.
# AISIEVAL_WORKBOOK overrides it if the workbook lives elsewhere.
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.environ.get(
    "AISIEVAL_WORKBOOK", os.path.join(_ROOT, "dataset", "AISIEVAL_V13.xlsx"))
SHEET = "AISIEVAL_V13"

# V13 = V12 with one adjudicated change: APOLLO-2026-07-ALI4 moved Tier A -> Tier C. Its finding is
# a process account of a red-team campaign ("recommendations, which Anthropic implemented"), not an
# empirical model finding, so it fails §4 gate 1. Response fields were cleared to match the Tier C
# convention observed on all 349 existing Tier C rows, and Finding Type moved capability-finding ->
# methodology because "capability-finding" contradicts Eval? = no.
# V12 remains untouched at ~/evaluating-the-evaluators/AISI  Eval Findings.xlsx, sheet AISIEVAL_V12.


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))          # Lag is stored as a float in V12; render it as an integer
    return str(v).strip()


def sheet():
    """The live worksheet, opened read-only-in-spirit. Never write through this handle."""
    return openpyxl.load_workbook(WORKBOOK, data_only=True)[SHEET]


def rows():
    """Every finding as a dict, keyed in sheet order."""
    ws = sheet()
    hdr = [norm(c.value) for c in ws[1]]
    out = []
    for r in range(2, ws.max_row + 1):
        d = {h: norm(ws.cell(r, i).value) for i, h in enumerate(hdr, 1) if h}
        if d.get("Finding ID"):
            out.append(d)
    return out
