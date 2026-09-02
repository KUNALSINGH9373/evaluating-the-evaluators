#!/usr/bin/env python3
"""15b — shortfall rate by year, on the HEADLINE metric.

Figure 15 plots the no-action rate, which trends the opposite way to the headline. Slide 7
describes the falls-short trend, so it needs a chart of that. Additive: fig 15 is untouched.
"""
import os, collections, datetime
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
import dataset_source
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palette import *   # single source of colour truth


OUT = CHARTS_OUT
plt.rcParams.update({"figure.dpi": 200, "savefig.dpi": 200, "font.family": "DejaVu Sans",
                     "figure.facecolor": "white", "axes.facecolor": "white",
                     "savefig.facecolor": "white"})
GREY_LINE = MUTED


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


P = dataset_source.WORKBOOK
ws = openpyxl.load_workbook(P, data_only=True)[dataset_source.SHEET]
hdr = [norm(c.value) for c in ws[1]]
R = [{h: norm(ws.cell(r, i).value) for i, h in enumerate(hdr, 1) if h} for r in range(2, ws.max_row + 1)]
R = [r for r in R if r.get("Finding ID")]
tier = lambda r: ("A" if r["Action Trackable?"] == "yes" else "B" if r["Eval? (trackable)"] == "yes" else "C")
H = [r for r in R if tier(r) == "A" and r["Severity (C1/C2) majority"] == "C1"]

years = sorted({r["Publication Date"][:4] for r in H if r["Publication Date"]})
n, short, noact = [], [], []
for y in years:
    g = [r for r in H if r["Publication Date"][:4] == y]
    n.append(len(g))
    short.append(sum(1 for r in g if r["Proportionality"] != "Proportionate") / len(g) * 100)
    noact.append(sum(1 for r in g if r["Proportionality"] == "Accountability gap (no action)") / len(g) * 100)

fig, ax = plt.subplots(figsize=(12, 7.4))
ax.plot(years, short, "-o", color=RED, linewidth=4.5, markersize=15, zorder=3,
        label="falls short of the standard  (headline metric)")
ax.plot(years, noact, "--s", color=GREY_LINE, linewidth=2.6, markersize=10, zorder=2,
        label="no response located at all")
for x, y, m in zip(years, short, n):
    ax.annotate(f"{y:.0f}%", (x, y), textcoords="offset points", xytext=(0, 20),
                ha="center", fontsize=26, color=RED, fontweight="bold")
    ax.annotate(f"n={m}", (x, y), textcoords="offset points", xytext=(0, -30),
                ha="center", fontsize=17, color=MUTED)
for x, y in zip(years, noact):
    ax.annotate(f"{y:.0f}%", (x, y), textcoords="offset points", xytext=(0, -30),
                ha="center", fontsize=19, color=GREY_LINE)
ax.set_ylim(0, 112)
ax.set_ylabel("share of significant-risk findings", fontsize=22)
ax.set_title("Shortfall is narrowing, but remains the norm", fontsize=31, color="#111111", pad=22)
ax.tick_params(labelsize=25)
ax.grid(axis="y", color=GRID, zorder=0)
for s in ("top", "right"):
    ax.spines[s].set_visible(False)
ax.legend(fontsize=17, frameon=False, loc="lower left")
fig.text(0.02, 0.015,
         f"Tier A ∩ severity C1, n={len(H)}. 2023 and 2024 are small samples; 2026 is partial (cutoff 29 August 2026). "
         "The two metrics trend\nin opposite directions: part of the decline in shortfall reflects composition — later "
         "years hold more pre-deployment\nevaluations, which are answered in the launch card by construction.",
         fontsize=13.5, color=INK_2, linespacing=1.6)
fig.tight_layout(rect=[0, 0.10, 1, 1])
p = os.path.join(OUT, "15b_shortfall_rate_by_year.png")
fig.savefig(p, bbox_inches="tight", pad_inches=0.3)
plt.close(fig)
print(f"wrote {p}")
print("  falls short:", " ".join(f"{y}={s:.0f}%" for y, s in zip(years, short)))
print("  no action  :", " ".join(f"{y}={s:.0f}%" for y, s in zip(years, noact)))
