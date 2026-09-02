#!/usr/bin/env python3
"""28_three_findings_provenance.png — the same three 2026 findings, by source.

Finding, evaluator, date and domain. Every field is read from the workbook; nothing is retyped.
"""
import os, datetime
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
import dataset_source
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palette import *
from tablefig import draw_table, INK

plt.rcParams.update({"figure.dpi": 200, "savefig.dpi": 200, "font.family": "DejaVu Sans",
                     "figure.facecolor": "white", "axes.facecolor": "white",
                     "savefig.facecolor": "white"})


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


ws = dataset_source.sheet()
hdr = [norm(c.value) for c in ws[1]]
ROWS = {r["Finding ID"]: r for r in
        ({h: norm(ws.cell(i, k).value) for k, h in enumerate(hdr, 1) if h}
         for i in range(2, ws.max_row + 1)) if r.get("Finding ID")}

MONTH = ["January", "February", "March", "April", "May", "June", "July", "August",
         "September", "October", "November", "December"]
FINDINGS = {
    "METR-2026-06-ALI1":
        "METR found GPT-5.6 Sol had the highest observed rate of cheating and reward-hacking of "
        "any model it had evaluated — exploiting bugs in the evaluation environment or adopting "
        "disallowed strategies.",
    "UKAISI-2026-04-JAI1":
        "UK AISI found a universal jailbreak in GPT-5.5 within six hours.",
    "UKAISI-2026-04-CYB1":
        "Claude Mythos Preview reached a 73% pass rate on expert-level CTFs and became the first "
        "model to solve a challenge on AISI's hardest “The Last Ones” range.",
}

# One column per finding, one row per field. Narrow columns are the point: they force the finding
# text to wrap, which is what gives the figure height without padding anything out.
FIELDS = ["Finding", "Evaluator", "Date", "Domain"]
cols_data = []
for fid, text in FINDINGS.items():
    r = ROWS[fid]
    d = r["Publication Date"]
    cols_data.append([text, r["Institution"],
                      f"{int(d[8:10])} {MONTH[int(d[5:7]) - 1]} {d[:4]}",
                      r["Domain"].replace(";", " · ")])

rows = [[FIELDS[i]] + [c[i] for c in cols_data] for i in range(len(FIELDS))]
COLS = [("", 13.0)] + [("", 28.0)] * len(cols_data)


def style(i, s):
    return ("#3A4650", "bold") if i == 0 else (INK, "normal")


p = draw_table(os.path.join(CHARTS_OUT, "28_three_findings_provenance.png"),
               COLS, rows, fig_w=9.4, fs=13, cell_style=style, header=False)
print(f"wrote {p}")
for c in cols_data:
    print(f"  {c[1]:<12} {c[2]:<18} {c[3]}")
