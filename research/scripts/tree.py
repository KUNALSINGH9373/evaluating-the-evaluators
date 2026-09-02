#!/usr/bin/env python3
"""13_institution_type_tree.png — institution tree: corpus → type → institution.

Owns this figure outright. It began as a top-down tree inside hero.py whose leaf labels collided,
and has kept the tree topology since: root, four branches, leaves.

What changed is everything around the structure. Group container boxes are gone — a rectangle per
branch is decoration, and it cost a great deal of blank space. The axes now fills the figure, so
the branches use the full width instead of two thirds of it, and every size is set to be read
from poster distance.

Leaves are proportional bars: bar LENGTH is the finding count on one scale shared across all four
branches, the institution NAME sits inside its bar where it fits, and the count is always
labelled outside the bar's end.
"""
import os, collections, datetime
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch
import openpyxl
import dataset_source
import sys; sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from palette import *   # single source of colour truth


OUT = CHARTS_OUT
plt.rcParams.update({"figure.dpi": 200, "savefig.dpi": 200, "font.family": "DejaVu Sans",
                     "figure.facecolor": "white", "axes.facecolor": "white",
                     "savefig.facecolor": "white"})
PRIM = ["Government", "Non-Profit (AIEF)", "Non-Profit (Independent)", "For-Profit"]
COL = INSTTYPE
TOPN = 5


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


ws = dataset_source.sheet()
hdr = [norm(c.value) for c in ws[1]]
R = [{h: norm(ws.cell(r, i).value) for i, h in enumerate(hdr, 1) if h}
     for r in range(2, ws.max_row + 1)]
R = [r for r in R if r.get("Finding ID")]


def prim(r):
    """A finding's primary institution type. Compound values fold to the first PRIM match, so
    every finding lands in exactly one group and the four groups sum to the corpus."""
    p = [x.strip() for x in r.get("Institution Type", "").split(";") if x.strip()]
    for k in PRIM:
        if k in p:
            return k
    return p[0] if p else "Other"


grp = collections.defaultdict(list)
for r in R:
    grp[prim(r)].append(r)
order = sorted(PRIM, key=lambda k: -len(grp[k]))
assert sum(len(grp[k]) for k in order) == len(R), "a finding fell outside the four groups"

groups = []
for k in order:
    top = [(short_inst(nm), v, False)
           for nm, v in collections.Counter(r["Institution"] for r in grp[k]).most_common(TOPN)]
    other = len(grp[k]) - sum(v for _, v, _ in top)
    leaves = top + ([("Other institutions", other, True)] if other > 0 else [])
    groups.append((k, len(grp[k]), leaves))

VMAX = max(v for _, _, leaves in groups for _, v, _ in leaves)
NLEAF = sum(len(lv) for _, _, lv in groups)

# ---- geometry, in row units; one row unit is UNIT_IN inches tall ---------------------------
UNIT_IN = 0.56
GAP = 0.95                    # between the last leaf of a branch and the first of the next
TITLE = 2.55                  # title block above the first heading
FOOT = 0.85                   # footnote below the last bar
BH = 0.74                     # bar height, in row units
RX, RW = 0.0, 11.5            # root node
SPINE = 13.5                  # vertical connector, root out to the branches
TX, TW = 15.5, 26.0           # branch node — wide enough for "Non-Profit (Independent)" at FS_BRANCH
LSPINE = 43.5                 # vertical connector, branch out to its leaves
BX, BW = 46.0, 46.0           # bars start here; BW is the length of a full-scale bar
FS_NAME, FS_NUM = 24, 26
FS_BIG = 32                   # the count line inside a root or branch node
FS_BRANCH = 24                # the branch's type name — sized to read, not as a caption
FS_SMALL = 19                 # "all findings", under the root count
FIG_W = 20.0
WIRE_C = "#C9D6E0"

total = TITLE + NLEAF + GAP * (len(groups) - 1) + FOOT
fig, ax = plt.subplots(figsize=(FIG_W, total * UNIT_IN))
# The axes fills the figure. Default margins left the bars using two thirds of the width while
# the oversized title overflowed to the right, and bbox_inches="tight" then padded the canvas out
# to fit it — which is where the empty right-hand band came from.
fig.subplots_adjust(left=0.010, right=0.997, top=0.997, bottom=0.006)
ax.set_xlim(0, 100)
ax.set_ylim(total, 0)         # inverted: y grows downward
ax.axis("off")

fig.canvas.draw()
_rend = fig.canvas.get_renderer()


def text_w(s, fs, weight="normal"):
    """Width of a string in data units — measured, not guessed, so the fits-inside test is real."""
    t = ax.text(0, 0, s, fontsize=fs, fontweight=weight, alpha=0)
    bb = t.get_window_extent(renderer=_rend)
    t.remove()
    inv = ax.transData.inverted()
    return abs(inv.transform((bb.width, 0))[0] - inv.transform((0, 0))[0])


ax.text(0, 0.80, "Who reports findings, by institution type",
        fontsize=48, color="#111111", fontweight="bold", va="baseline", ha="left")
# Kept short on purpose: a full sentence at this weight runs wider than the figure.
ax.text(0, 1.78, f"Bar length = findings, on one scale across all four branches  ·  "
        f"top {TOPN} institutions per type",
        fontsize=26, color="#4A4A4A", fontweight="bold", va="baseline", ha="left")

# vertical extent of each branch's block of leaves
y, blocks = TITLE, []
for k, cnt, leaves in groups:
    blocks.append((k, cnt, leaves, y, y + len(leaves)))
    y += len(leaves) + GAP

mids = [(a + b) / 2 for _, _, _, a, b in blocks]
root_mid = (blocks[0][3] + blocks[-1][4]) / 2

# ---- root -------------------------------------------------------------------------------
ax.add_patch(FancyBboxPatch((RX, root_mid - 1.35), RW, 2.7,
                            boxstyle="round,pad=0,rounding_size=0.26",
                            facecolor="#EDF2F6", edgecolor=WIRE_C, linewidth=1.6))
ax.text(RX + RW / 2, root_mid - 0.28, f"{len(R):,}", ha="center", va="center",
        fontsize=FS_BIG, color="#111111", fontweight="bold")
ax.text(RX + RW / 2, root_mid + 0.72, "all findings", ha="center", va="center",
        fontsize=FS_SMALL, color=MUTED)
ax.plot([SPINE, SPINE], [min(mids), max(mids)], color=WIRE_C, lw=2.2)
ax.plot([RX + RW, SPINE], [root_mid, root_mid], color=WIRE_C, lw=2.2)

# ---- one branch per institution type ------------------------------------------------------
for (k, cnt, leaves, y0, y1), mid in zip(blocks, mids):
    col, ink = COL[k], shade(COL[k], 0.30)
    ax.plot([SPINE, TX], [mid, mid], color=col, lw=2.6)
    ax.add_patch(FancyBboxPatch((TX, mid - 1.55), TW, 3.1,
                                boxstyle="round,pad=0,rounding_size=0.26",
                                facecolor=col, edgecolor="none"))
    ax.text(TX + 1.2, mid - 0.40, f"{cnt}", ha="left", va="center",
            fontsize=FS_BIG, color="white", fontweight="bold")
    ax.text(TX + 1.2, mid + 0.80, k, ha="left", va="center",
            fontsize=FS_BRANCH, color="white")
    assert text_w(k, FS_BRANCH) + 2.4 <= TW, f"branch label overflows its node: {k}"

    lys = [y0 + i + 0.5 for i in range(len(leaves))]
    ax.plot([LSPINE, LSPINE], [min(lys), max(lys)], color=col, lw=2.0, alpha=0.5)
    ax.plot([TX + TW, LSPINE], [mid, mid], color=col, lw=2.0, alpha=0.5)

    # a name goes inside its bar only if it fits AND every longer bar in this branch does too,
    # otherwise a shorter bar ends up looking more prominent than a longer one
    inside, still_fitting = [], True
    for nm, v, _ in leaves:
        still_fitting = (still_fitting
                         and text_w(nm, FS_NAME, "bold") + 1.6 <= max(v / VMAX * BW, 0.6))
        inside.append(still_fitting)

    for (nm, v, is_other), nm_inside, ly in zip(leaves, inside, lys):
        w = max(v / VMAX * BW, 0.6)          # floor so a 4-finding bar is still visible
        ax.plot([LSPINE, BX], [ly, ly], color=col, lw=2.0, alpha=0.5)
        ax.add_patch(FancyBboxPatch((BX, ly - BH / 2), w, BH,
                                    boxstyle="round,pad=0,rounding_size=0.14",
                                    facecolor=col, edgecolor="none",
                                    alpha=0.5 if is_other else 1.0))
        ax.text(BX + w + 0.9, ly, f"{v}", ha="left", va="center",
                fontsize=FS_NUM, color=ink, fontweight="bold")
        if nm_inside:
            ax.text(BX + 0.9, ly, nm, ha="left", va="center", fontsize=FS_NAME,
                    color="white", fontweight="bold")
        else:
            ax.text(BX + w + 0.9 + text_w(str(v), FS_NUM, "bold") + 1.3, ly, nm,
                    ha="left", va="center", fontsize=FS_NAME, color="#1A1A1A")

ax.text(0, total - 0.15,
        f'Institution Type field; compound values (e.g. "Government;Lab") fold into their primary '
        f'type, so the four branches sum to {len(R):,}.',
        fontsize=16, color="#4A4A4A", va="baseline", ha="left")

p = os.path.join(OUT, "13_institution_type_tree.png")
fig.savefig(p, bbox_inches="tight", pad_inches=0.22)
plt.close(fig)
print(f"wrote {p}  ({NLEAF} bars, scale max {VMAX}, {FIG_W:.1f}x{total * UNIT_IN:.1f} in)")
for k, cnt, leaves in groups:
    print(f"  {k:<26} {cnt:>4}   " + " · ".join(f"{nm} {v}" for nm, v, _ in leaves))

# The old bar version of this chart is superseded by this figure; keep it from reappearing.
old = os.path.join(OUT, "13_institution_type.png")
if os.path.exists(old):
    os.remove(old)
    print("  removed superseded 13_institution_type.png")
