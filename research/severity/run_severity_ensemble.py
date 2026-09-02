#!/usr/bin/env python3
"""
run_severity_ensemble.py — Stage 5 severity coding for "Evaluating the Evaluators".

ONE frozen prompt (severity_prompt.txt) is sent verbatim and identically to three
models. Majority vote wins; all three votes are recorded and never overwritten.

    Claude Sonnet 5   ->  column "Sonnet5 vote"
    GPT-5.5           ->  column "GPT-5.5 vote"
    Gemini 3.1 Pro    ->  column "Gemini3.1 vote"
    majority          ->  column "Severity (C1/C2) majority"

Values are the literal strings "C1", "C2", or "ERR" — matching v10.

Modes
-----
  --calibrate N   Re-code N random rows that ALREADY have v10 votes and report
                  agreement with them. RUN THIS FIRST. See "Provenance" below.
  --run           Code every unvoted row of the target workbook.
  --rows ID,ID    Code only these Finding IDs (use for re-votes).
  --dry-run       Build payloads and print one, call nothing, spend nothing.

Every model reply is appended to severity_votes.jsonl before anything is written
back, so a crash or a rate-limit never loses paid work. Re-running skips rows
already present in the JSONL unless --force is given.

DETERMINISM — the ensemble is no longer temperature-0
-----------------------------------------------------
Claude Sonnet 5 rejects `temperature` outright ("deprecated for this model") and
GPT-5.5 accepts only its default of 1. Both are therefore called WITHOUT a
temperature parameter; only Gemini still runs at temperature 0. OpenAI gets a
fixed seed (20260815) for best-effort repeatability; Anthropic exposes no such
knob.

Consequence: re-running the same row can produce a different vote from those two
models. Votes are written once and never overwritten, and every reply is kept in
the journal, so the record is fixed even though the process is not reproducible
bit-for-bit. Report this in the methods section — v10's votes may have been
taken at temperature 0, which is a real difference in sampling regime between
the two halves of the dataset, independent of the prompt-reconstruction question
below.

Provenance — read this before trusting a number
-----------------------------------------------
The original frozen prompt file was not on disk when v11 was coded. The prompt in
severity_prompt.txt is RECONSTRUCTED verbatim from v10_METHODOLOGY.md §9.2-9.5,
which quotes the rubric and the four critical rules in full. It is faithful to the
documented spec but it is not byte-identical to whatever produced v10's 456 rows.

Therefore: --calibrate is not optional. It re-codes a sample of already-voted v10
rows and reports per-model agreement. High agreement (>= ~0.90) means the
reconstruction reproduces the original behaviour and v11 votes are comparable to
v10. Low agreement means you must either re-code all of v10 with this prompt or
report the discontinuity in the paper. Do not code v11 and merge it into v10
without running this first.

Environment
-----------
  ANTHROPIC_API_KEY   required
  OPENAI_API_KEY      required
  Gemini, either:
    GOOGLE_APPLICATION_CREDENTIALS  service-account JSON (project_id read from it)
    GCP_LOCATION                    default "global"
  or:
    GOOGLE_API_KEY / GEMINI_API_KEY  AI Studio key instead of Vertex
  SEV_MODEL_ANTHROPIC / SEV_MODEL_OPENAI / SEV_MODEL_GOOGLE
                      optional model-id overrides; defaults below are the
                      intended models but VERIFY them against the current API
                      before a production run.
"""

from __future__ import annotations

import argparse
import json
import os
import random
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import urllib.error
import urllib.request

HERE = Path(__file__).resolve().parent
PROMPT_FILE = HERE / "severity_prompt.txt"          # v1.1, reassigned by --prompt
PROMPT_VERSION = "1.1"                              # recorded on every vote; kept in lockstep
PROMPT_BY_VERSION = {"1.0": HERE / "severity_prompt_v1.0_FROZEN.txt",
                     "1.1": HERE / "severity_prompt.txt"}
JOURNAL = HERE / "severity_votes.jsonl"
V10 = HERE / "v10.csv"
# The dataset has exactly one source of truth; every other reader imports it. Pointing this at a
# stale path once wrote votes at a file that no longer exists — and would silently target the
# wrong workbook if one did.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
import dataset_source as _ds
DEFAULT_TARGET = Path(_ds.WORKBOOK)
TARGET_SHEET = _ds.SHEET

COL_MAJORITY = "Severity (C1/C2) majority"
COL_VOTE = {
    "sonnet5": "Sonnet5 vote",
    "gpt55": "GPT-5.5 vote",
    "gemini31": "Gemini3.1 vote",
}

MODELS = {
    "sonnet5": os.environ.get("SEV_MODEL_ANTHROPIC", "claude-sonnet-5"),
    "gpt55": os.environ.get("SEV_MODEL_OPENAI", "gpt-5.5"),
    "gemini31": os.environ.get("SEV_MODEL_GOOGLE", "gemini-3.1-pro-preview"),
}

MAX_RETRIES = 5
BASE_BACKOFF = 4.0
TIMEOUT = 120


# --------------------------------------------------------------------------
# the one prompt
# --------------------------------------------------------------------------

def load_prompt() -> tuple[str, str]:
    """Split severity_prompt.txt into (system, user_template) on its own markers."""
    raw = PROMPT_FILE.read_text(encoding="utf-8")
    try:
        _, rest = raw.split("SYSTEM\n" + "=" * 80, 1)
        system, user = rest.split("USER MESSAGE TEMPLATE\n" + "=" * 80, 1)
    except ValueError:
        sys.exit(f"{PROMPT_FILE.name}: section markers not found — file was edited.")
    system = system.rsplit("=" * 80, 1)[0].strip()
    return system, user.strip()


def build_user(template: str, row: dict) -> str:
    return template.format(
        finding_id=row.get("Finding ID", ""),
        domain=row.get("Domain", "") or "unspecified",
        models=row.get("Models / Systems", "") or "unspecified",
        institution=row.get("Institution", "") or "unspecified",
        finding=(row.get("Finding", "") or "").strip(),
        finding_quote=(row.get("Finding Quote", "") or "").strip() or "(none recorded)",
    )


# --------------------------------------------------------------------------
# transport — identical prompt, three providers
# --------------------------------------------------------------------------

class AuthError(RuntimeError):
    """Credentials or model id are wrong. Never worth retrying."""


class QuotaError(RuntimeError):
    """Out of credit / hard quota. Stop the whole run - retrying only wastes time."""


QUOTA_MARKERS = ("insufficient_quota", "insufficient credit", "billing", "exceeded your current quota",
                 "credit balance is too low", "quota exceeded", "RESOURCE_EXHAUSTED",
                 "payment required", "account is not active")


def _post(url: str, payload: dict, headers: dict, form: str | None = None) -> dict:
    if form is not None:
        body = form.encode()
        headers = {"Content-Type": "application/x-www-form-urlencoded", **headers}
    else:
        body = json.dumps(payload).encode()
        headers = {"Content-Type": "application/json", **headers}
    req = urllib.request.Request(url, data=body, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        detail = ""
        try:
            detail = e.read().decode()[:300]
        except Exception:  # noqa: BLE001
            pass
        low = detail.lower()
        if any(m.lower() in low for m in QUOTA_MARKERS) or e.code == 402:
            raise QuotaError(f"HTTP {e.code} — {detail}") from e
        if e.code in (400, 401, 403, 404):
            raise AuthError(f"HTTP {e.code} — {detail}") from e
        raise


def call_anthropic(system: str, user: str) -> str:
    key = os.environ["ANTHROPIC_API_KEY"]
    out = _post(
        "https://api.anthropic.com/v1/messages",
        {
            "model": MODELS["sonnet5"],
            "max_tokens": 1024,
            # temperature omitted: deprecated for Claude Sonnet 5. See DETERMINISM note below.
            "system": system,
            "messages": [{"role": "user", "content": user}],
        },
        {"x-api-key": key, "anthropic-version": "2023-06-01"},
    )
    return "".join(b.get("text", "") for b in out.get("content", []))


def call_openai(system: str, user: str) -> str:
    key = os.environ["OPENAI_API_KEY"]
    out = _post(
        "https://api.openai.com/v1/chat/completions",
        {
            "model": MODELS["gpt55"],
            # temperature omitted: GPT-5.5 accepts only its default. Fixed seed instead.
            "seed": 20260815,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
        },
        {"Authorization": f"Bearer {key}"},
    )
    return out["choices"][0]["message"]["content"]


def _project_from_sa() -> str:
    """Read project_id straight out of the service-account JSON, so it is never guessed."""
    sa = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")
    if sa and os.path.exists(sa):
        try:
            return json.load(open(sa)).get("project_id", "")
        except Exception:  # noqa: BLE001
            return ""
    return ""


# GCP_PROJECT / GCP_LOCATION are the names used by the original runner; accept both.
VERTEX_PROJECT = (os.environ.get("SEV_VERTEX_PROJECT")
                  or os.environ.get("GCP_PROJECT")
                  or _project_from_sa())
VERTEX_LOCATION = (os.environ.get("SEV_VERTEX_LOCATION")
                   or os.environ.get("GCP_LOCATION") or "global")


def vertex_token() -> str:
    """Bearer token for Vertex, in order of preference.

    1. GOOGLE_ACCESS_TOKEN set directly (paste from any machine that has gcloud)
    2. `gcloud auth print-access-token` if the SDK is installed
    3. service-account JSON at GOOGLE_APPLICATION_CREDENTIALS, signed with openssl
    """
    tok = os.environ.get("GOOGLE_ACCESS_TOKEN", "").strip()
    if tok:
        return tok
    import shutil as _sh
    import subprocess as _sp

    if _sh.which("gcloud"):
        r = _sp.run(["gcloud", "auth", "print-access-token"], capture_output=True, text=True)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout.strip()
    sa = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")
    if sa and os.path.exists(sa):
        return _sa_token(sa)
    raise AuthError(
        "no Vertex credential. Set GOOGLE_ACCESS_TOKEN, or install gcloud and run "
        "`gcloud auth application-default login`, or point GOOGLE_APPLICATION_CREDENTIALS "
        "at a service-account JSON."
    )


def _sa_token(path: str) -> str:
    """Mint an access token from a service-account JSON using openssl for the RS256 signature."""
    import base64
    import subprocess as _sp
    import tempfile

    info = json.load(open(path))
    now = int(time.time())
    b64 = lambda d: base64.urlsafe_b64encode(json.dumps(d, separators=(",", ":")).encode()).rstrip(b"=")
    head = b64({"alg": "RS256", "typ": "JWT"})
    body = b64({
        "iss": info["client_email"],
        "scope": "https://www.googleapis.com/auth/cloud-platform",
        "aud": "https://oauth2.googleapis.com/token",
        "iat": now, "exp": now + 3600,
    })
    signing_input = head + b"." + body
    with tempfile.NamedTemporaryFile("w", suffix=".pem", delete=False) as f:
        f.write(info["private_key"]); pem = f.name
    try:
        sig = _sp.run(["openssl", "dgst", "-sha256", "-sign", pem],
                      input=signing_input, capture_output=True).stdout
    finally:
        os.unlink(pem)
    jwt = signing_input + b"." + base64.urlsafe_b64encode(sig).rstrip(b"=")
    out = _post("https://oauth2.googleapis.com/token", None, {}, form=(
        "grant_type=urn:ietf:params:oauth:grant-type:jwt-bearer&assertion=" + jwt.decode()))
    return out["access_token"]


def call_google(system: str, user: str) -> str:
    model = MODELS["gemini31"]
    payload = {
        "systemInstruction": {"parts": [{"text": system}]},
        "contents": [{"role": "user", "parts": [{"text": user}]}],
        "generationConfig": {"temperature": 0, "responseMimeType": "application/json"},
    }
    if VERTEX_PROJECT:
        host = ("aiplatform.googleapis.com" if VERTEX_LOCATION == "global"
                else f"{VERTEX_LOCATION}-aiplatform.googleapis.com")
        url = (f"https://{host}/v1/projects/{VERTEX_PROJECT}/locations/{VERTEX_LOCATION}"
               f"/publishers/google/models/{model}:generateContent")
        out = _post(url, payload, {"Authorization": "Bearer " + vertex_token()})
    else:
        key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY", "")
        out = _post(
            f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}",
            payload, {})
    parts = out["candidates"][0]["content"]["parts"]
    return "".join(p.get("text", "") for p in parts)


CALLERS = {"sonnet5": call_anthropic, "gpt55": call_openai, "gemini31": call_google}


# --------------------------------------------------------------------------
# parsing + the category rule
# --------------------------------------------------------------------------

def parse_vote(text: str) -> dict:
    """Extract the JSON object and RECOMPUTE category from the domain flags.

    The prompt tells the model category = 1 if any d* is true. We never trust it
    to have done that arithmetic — the flags are the finding of record.
    """
    m = re.search(r"\{.*\}", text, re.S)
    if not m:
        raise ValueError(f"no JSON object in reply: {text[:200]!r}")
    obj = json.loads(m.group(0))
    flags = {f"d{i}": bool(obj.get(f"d{i}", False)) for i in range(1, 9)}   # D8 added in prompt v1.1
    category = 1 if any(flags.values()) else 2
    if obj.get("category") not in (1, 2, "1", "2"):
        obj["_category_missing"] = True
    elif int(obj["category"]) != category:
        obj["_category_disagreed_with_flags"] = int(obj["category"])
    return {
        **flags,
        "category": category,
        "label": f"C{category}",
        "quote": (obj.get("quote") or "")[:1000],
        "confidence": obj.get("confidence"),
        "reason": (obj.get("reason") or "")[:400],
        "_flagged": {k: v for k, v in obj.items() if k.startswith("_")},
    }


def vote_one(key: str, system: str, user: str) -> dict:
    last = None
    for attempt in range(MAX_RETRIES):
        try:
            return parse_vote(CALLERS[key](system, user))
        except (AuthError, QuotaError):
            # Bad key, bad model id, or out of credit — retrying cannot fix any of them.
            raise
        except urllib.error.HTTPError as e:
            last = f"HTTP {e.code}"
            if e.code in (429, 500, 502, 503, 529):
                time.sleep(BASE_BACKOFF * (2 ** attempt))
                continue
            break
        except Exception as e:  # noqa: BLE001 — record and retry, never crash the run
            last = f"{type(e).__name__}: {e}"
            time.sleep(BASE_BACKOFF * (2 ** attempt))
    return {"label": "ERR", "error": last}


PLACEHOLDER = re.compile(r"\.\.\.|^$|^your|^sk-\.\.\.|^<")


def preflight(system: str, template: str, sample: dict) -> None:
    """One real call per provider before the batch. Abort loudly on auth failure.

    A whole run of ERR is worse than no run: it looks like data and it overwrites
    the sheet. Catch it on call 1, not call 1,689.
    """
    keys = {
        "sonnet5": os.environ.get("ANTHROPIC_API_KEY", ""),
        "gpt55": os.environ.get("OPENAI_API_KEY", ""),
        "gemini31": os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY", ""),
    }
    if VERTEX_PROJECT:
        keys.pop("gemini31", None)
    bad = [k for k, v in keys.items() if PLACEHOLDER.match(v.strip()) or len(v.strip()) < 20]
    if bad:
        sys.exit("These keys look like placeholders, not real credentials: "
                 + ", ".join(bad)
                 + "\nYou pasted the example text rather than your key. Nothing was called.")

    user = build_user(template, sample)
    fails = []
    print("preflight: one call per provider ...")
    for k in CALLERS:
        try:
            CALLERS[k](system, user)
            print(f"  {k:<10} {MODELS[k]:<22} ok")
        except AuthError as e:
            fails.append(f"  {k:<10} {MODELS[k]:<22} {e}")
        except Exception as e:  # noqa: BLE001
            fails.append(f"  {k:<10} {MODELS[k]:<22} {type(e).__name__}: {e}")
    if fails:
        sys.exit("\npreflight FAILED — nothing was coded, nothing was written:\n"
                 + "\n".join(fails)
                 + "\n\nCheck the API key and the model id for each failing provider."
                   "\nModel ids are overridable: SEV_MODEL_ANTHROPIC / SEV_MODEL_OPENAI / SEV_MODEL_GOOGLE")
    print()


def majority(labels: list[str]) -> str:
    real = [x for x in labels if x in ("C1", "C2")]
    if not real:
        return "ERR"
    return "C1" if real.count("C1") * 2 > len(real) else ("C2" if real.count("C2") * 2 > len(real) else "ERR")


# --------------------------------------------------------------------------
# Fleiss' kappa
# --------------------------------------------------------------------------

def fleiss_kappa(rows: list[list[str]]) -> float | None:
    """rows = list of per-item label lists. Items with any ERR are dropped."""
    items = [r for r in rows if all(x in ("C1", "C2") for x in r) and len(r) > 1]
    if not items:
        return None
    n = len(items[0])
    if any(len(r) != n for r in items):
        return None
    cats = ("C1", "C2")
    p_j = [sum(r.count(c) for r in items) / (len(items) * n) for c in cats]
    p_i = [(sum(r.count(c) ** 2 for c in cats) - n) / (n * (n - 1)) for r in items]
    p_bar = sum(p_i) / len(items)
    pe = sum(p * p for p in p_j)
    return None if pe == 1 else (p_bar - pe) / (1 - pe)


# --------------------------------------------------------------------------
# journal
# --------------------------------------------------------------------------

def journal_read() -> dict:
    """Replay the journal into {finding_id: {model_key: vote}}.

    Only SUCCESSFUL legs are kept. An ERR leg is treated as never attempted, so a
    run that died on a balance error resumes and retries exactly those legs
    instead of skipping them as done.
    """
    if not JOURNAL.exists():
        return {}
    out: dict = {}
    for line in JOURNAL.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
        except json.JSONDecodeError:
            continue
        fid = rec.get("finding_id")
        if not fid:
            continue
        slot = out.setdefault(fid, {})
        for k, v in (rec.get("votes") or {}).items():
            if isinstance(v, dict) and v.get("label") in ("C1", "C2"):
                slot[k] = v
    return out


def complete(votes: dict) -> bool:
    """All three legs answered."""
    return all(k in votes for k in CALLERS)


def decidable(votes: dict) -> bool:
    """Enough legs to carry a majority. Two agreeing votes decide a row 2-0."""
    return len([k for k in CALLERS if k in votes]) >= 2


def failed_attempts(fid: str) -> dict:
    """How many times each leg has already ERRed for this row, from the journal."""
    n: dict = {}
    if not JOURNAL.exists():
        return n
    for line in JOURNAL.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
        except json.JSONDecodeError:
            continue
        if rec.get("finding_id") != fid:
            continue
        for k, v in (rec.get("votes") or {}).items():
            if isinstance(v, dict) and v.get("label") == "ERR":
                n[k] = n.get(k, 0) + 1
    return n


def journal_append(rec: dict) -> None:
    with JOURNAL.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# --------------------------------------------------------------------------
# workbook / csv io
# --------------------------------------------------------------------------

def load_xlsx(path: Path):
    import openpyxl

    wb = openpyxl.load_workbook(path)
    # Name the sheet explicitly. sheetnames[0] silently targets whatever happens to be first,
    # which is how a vote could land in an archived sheet and be lost.
    if TARGET_SHEET in wb.sheetnames:
        ws = wb[TARGET_SHEET]
    elif len(wb.sheetnames) == 1:
        ws = wb[wb.sheetnames[0]]
    else:
        sys.exit(f"{path.name}: expected sheet {TARGET_SHEET!r}, found {wb.sheetnames}")
    hdr = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(hdr) if h}
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({h: ws.cell(r, i).value for h, i in idx.items()} | {"_row": r})
    return wb, ws, idx, rows


def load_v10_sample(n: int) -> list[dict]:
    import csv

    rows = [r for r in csv.DictReader(V10.open(encoding="utf-8", errors="ignore"))
            if r.get(COL_MAJORITY) in ("C1", "C2")]
    random.seed(20260815)
    return random.sample(rows, min(n, len(rows)))


# --------------------------------------------------------------------------
# driver
# --------------------------------------------------------------------------

def code_rows(rows: list[dict], system: str, template: str, force: bool, workers: int) -> dict:
    """Code rows leg-by-leg, resuming from whatever the journal already holds.

    Resume is per (row, model): a row whose Sonnet and GPT legs succeeded before a
    crash only re-calls Gemini. A QuotaError anywhere stops the run at once and
    everything already earned stays in the journal.
    """
    cached = {} if force else journal_read()
    def settled(fid):
        v = cached.get(fid, {})
        if complete(v):
            return True
        pr = failed_attempts(fid)
        # decided 2-0 and the third leg has repeatedly refused -> nothing left to try
        return decidable(v) and all(pr.get(k, 0) >= 2 for k in CALLERS if k not in v)
    todo = [r for r in rows if not settled(r.get("Finding ID"))]
    legs = sum(len(CALLERS) - len(cached.get(r.get("Finding ID"), {})) for r in todo)
    done_rows = len(rows) - len(todo)
    reused = sum(len(cached.get(r.get("Finding ID"), {})) for r in todo)
    print(f"{len(rows)} rows · {done_rows} already complete in journal · {len(todo)} to code")
    if reused:
        print(f"  resuming: {reused} individual model votes reused from a previous run")
    print(f"  {legs} calls to make")

    stop = {"quota": None}

    def work(row: dict) -> dict:
        fid = row.get("Finding ID")
        votes = dict(cached.get(fid, {}))
        if stop["quota"]:
            return None
        user = build_user(template, row)
        prior = failed_attempts(fid)
        need = [k for k in CALLERS if k not in votes and prior.get(k, 0) < 2]
        giveup = [k for k in CALLERS if k not in votes and prior.get(k, 0) >= 2]
        for k in giveup:
            votes[k] = {"label": "ERR", "error": "persistent: %d failed attempts, not retried" % prior[k]}
        with ThreadPoolExecutor(max(1, len(need))) as ex:
            futs = {k: ex.submit(vote_one, k, system, user) for k in need}
            for k, f in futs.items():
                try:
                    votes[k] = f.result()
                except QuotaError as e:
                    stop["quota"] = f"{k}: {e}"
                    votes[k] = {"label": "ERR", "error": str(e)}
                except AuthError as e:
                    stop["quota"] = f"{k}: {e}"
                    votes[k] = {"label": "ERR", "error": str(e)}
        labels = [votes.get(k, {}).get("label", "ERR") for k in ("sonnet5", "gpt55", "gemini31")]
        rec = {"finding_id": fid, "majority": majority(labels), "votes": votes,
               "prompt_version": PROMPT_VERSION, "models": dict(MODELS)}
        journal_append(rec)          # written before anything else can fail
        return rec

    want = {r.get("Finding ID") for r in rows}
    results = {fid: {"finding_id": fid,
                     "majority": majority([v.get(k, {}).get("label", "ERR")
                                           for k in ("sonnet5", "gpt55", "gemini31")]),
                     "votes": v}
               for fid, v in cached.items() if fid in want and decidable(v)}
    if todo:
        with ThreadPoolExecutor(workers) as ex:
            for i, rec in enumerate(ex.map(work, todo), 1):
                if rec:
                    results[rec["finding_id"]] = rec
                if i % 25 == 0 or i == len(todo):
                    print(f"  {i}/{len(todo)}")
                if stop["quota"]:
                    break
    if stop["quota"]:
        print("\n" + "=" * 72)
        print("RUN STOPPED — out of credit or credentials rejected:")
        print("  " + stop["quota"])
        print(f"\nEverything completed so far is saved in {JOURNAL.name}.")
        print("Top up the account and re-run the SAME command — it will resume from")
        print("exactly the legs that are missing and will not re-pay for finished ones.")
        print("=" * 72)
    return results


def report(results: dict) -> None:
    import collections

    maj = collections.Counter(r["majority"] for r in results.values())
    print("\nmajority:", dict(maj))
    for k, col in COL_VOTE.items():
        c = collections.Counter(r["votes"][k]["label"] for r in results.values() if "votes" in r)
        print(f"  {col:<22} {dict(c)}")
    grids = [[r["votes"][k]["label"] for k in ("sonnet5", "gpt55", "gemini31")]
             for r in results.values() if "votes" in r]
    k = fleiss_kappa(grids)
    print(f"\nFleiss' kappa: {k:.4f}" if k is not None else "\nFleiss' kappa: n/a")
    split = [r["finding_id"] for r, g in zip(results.values(), grids) if len(set(g)) > 1]
    print(f"non-unanimous: {len(split)}")
    for fid in split[:25]:
        print("   ", fid)
    bad = [r["finding_id"] for r in results.values()
           if any(v.get("_flagged") for v in r.get("votes", {}).values())]
    if bad:
        print(f"\nmodels whose stated category disagreed with their own flags "
              f"(flags won): {len(bad)}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    ap.add_argument("--calibrate", type=int, metavar="N",
                    help="re-code N already-voted v10 rows and report agreement")
    ap.add_argument("--run", action="store_true")
    ap.add_argument("--rows", help="comma-separated Finding IDs")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--check", action="store_true",
                    help="verify all three API keys and model ids with one real call each, then exit")
    ap.add_argument("--prompt", choices=("1.0", "1.1"), default="1.1",
                    help="which frozen prompt to send. 1.0 = the rubric that coded all 1001 V12 "
                         "rows (D1-D7); 1.1 adds D8 acute individual harm. Votes are tagged with "
                         "this version, so the two arms never pool by accident.")
    ap.add_argument("--force", action="store_true", help="ignore the journal")
    ap.add_argument("--workers", type=int, default=4)
    a = ap.parse_args()

    global PROMPT_FILE, PROMPT_VERSION
    PROMPT_FILE, PROMPT_VERSION = PROMPT_BY_VERSION[a.prompt], a.prompt
    if not PROMPT_FILE.exists():
        sys.exit(f"prompt v{a.prompt} not found at {PROMPT_FILE}")

    system, template = load_prompt()
    print(f"prompt v{PROMPT_VERSION}  ({PROMPT_FILE.name})", file=sys.stderr)

    if a.check:
        print("environment:")
        for k in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY"):
            v = os.environ.get(k, "").strip()
            print("  %-32s %s" % (k, ("set, %d chars, starts %r" % (len(v), v[:7])) if v else "NOT SET"))
        if VERTEX_PROJECT:
            print("  %-32s %s" % ("GOOGLE_APPLICATION_CREDENTIALS",
                                  os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "(not set)")))
            print("  %-32s %s" % ("  -> project_id", VERTEX_PROJECT))
            print("  %-32s %s" % ("  -> location", VERTEX_LOCATION))
            if os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY"):
                print("  note: GOOGLE_API_KEY/GEMINI_API_KEY is set but IGNORED - Vertex is active.")
        else:
            v = (os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY", "")).strip()
            print("  %-32s %s" % ("GOOGLE_API_KEY",
                                  ("set, %d chars, starts %r" % (len(v), v[:7])) if v else "NOT SET"))
            print("  (no Vertex config: set GOOGLE_APPLICATION_CREDENTIALS for project witt-c9-gemini)")
        print()
        _, _, _, rows = load_xlsx(a.target)
        preflight(system, template, rows[0])       # exits non-zero on any failure
        print("All three providers answered. Models used:")
        for k, m in MODELS.items():
            print("  %-10s %s" % (k, m))
        print("\nNext:  python3 %s --calibrate 60" % os.path.basename(__file__))
        return

    if a.calibrate:
        rows = load_v10_sample(a.calibrate)
    else:
        _, _, _, rows = load_xlsx(a.target)
        if a.rows:
            want = {x.strip() for x in a.rows.split(",")}
            rows = [r for r in rows if r.get("Finding ID") in want]
        elif a.run:
            rows = [r for r in rows if r.get(COL_MAJORITY) in (None, "", "ERR")]
        else:
            sys.exit("pick a mode: --calibrate N | --run | --rows ID,ID")

    if not rows:
        sys.exit("nothing to do")

    if a.dry_run:
        print("=== SYSTEM ===\n" + system[:1500] + "\n...\n")
        print("=== USER (row 1) ===\n" + build_user(template, rows[0]))
        print(f"\n{len(rows)} rows x 3 models = {len(rows) * 3} calls")
        return

    for var in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY"):
        if not os.environ.get(var):
            sys.exit(f"{var} is not set")
    if not (os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY") or VERTEX_PROJECT):
        sys.exit("Set GOOGLE_API_KEY, or SEV_VERTEX_PROJECT for Vertex AI")

    # Calibration codes v10 rows, production codes v11 rows. They must not share
    # a journal keyed by Finding ID, or one silently resumes into the other.
    global JOURNAL
    JOURNAL = HERE / ("severity_votes_calibrate.jsonl" if a.calibrate else "severity_votes.jsonl")

    preflight(system, template, rows[0])
    results = code_rows(rows, system, template, a.force, a.workers)
    report(results)

    if a.calibrate:
        agree = {k: [0, 0] for k in list(COL_VOTE) + ["majority"]}
        for row in rows:
            rec = results.get(row["Finding ID"])
            if not rec:
                continue
            for k, col in COL_VOTE.items():
                if row.get(col) in ("C1", "C2"):
                    agree[k][1] += 1
                    agree[k][0] += rec["votes"][k]["label"] == row[col]
            agree["majority"][1] += 1
            agree["majority"][0] += rec["majority"] == row[COL_MAJORITY]
        errs = sum(1 for r in results.values() if r["majority"] == "ERR")
        if errs:
            print(f"\n{errs}/{len(results)} rows returned ERR. Agreement below is "
                  f"meaningless until those are fixed — an error is not a disagreement.")
        print("\n=== agreement with the stored v10 votes ===")
        for k, (hit, tot) in agree.items():
            if tot:
                print(f"  {k:<10} {hit}/{tot} = {hit / tot:.3f}")
        m = agree["majority"]
        if errs:
            pass
        elif m[1] and m[0] / m[1] < 0.90:
            print("\n  Majority agreement is below 0.90. The reconstructed prompt does NOT\n"
                  "  reproduce the original coding. Do not merge v11 severities into v10\n"
                  "  without either re-coding all of v10 or reporting the discontinuity.")
        print("\n(calibration only — nothing was written back)")
        return

    # Refuse to overwrite the sheet with a wall of ERR. A failed run must leave
    # no trace in the data — errors belong in the journal, not in the dataset.
    err = sum(1 for r in results.values() if r["majority"] == "ERR")
    if results and err / len(results) > 0.10:
        sys.exit(f"\nABORTED: {err}/{len(results)} rows came back ERR "
                 f"({err / len(results):.0%}). Nothing was written to {a.target.name}.\n"
                 f"Diagnose with:  grep -o '\"error\": \"[^\"]*\"' {JOURNAL.name} | sort | uniq -c\n"
                 f"Fix the cause, delete {JOURNAL.name}, and re-run.")

    wb, ws, idx, sheet_rows = load_xlsx(a.target)
    written = skipped = 0
    for row in sheet_rows:
        rec = results.get(row.get("Finding ID"))
        if not rec:
            continue
        if rec["majority"] == "ERR":
            skipped += 1  # leave the cell blank so the row stays visibly uncoded
            continue
        # a leg that never answered is recorded as ERR alongside the two that did,
        # so a 2-0 decision is visible in the vote columns rather than hidden
        ws.cell(row["_row"], idx[COL_MAJORITY]).value = rec["majority"]
        for k, col in COL_VOTE.items():
            ws.cell(row["_row"], idx[col]).value = rec["votes"][k]["label"]
        written += 1
    wb.save(a.target)
    print(f"\nwrote {written} rows to {a.target}")
    if skipped:
        print(f"left {skipped} ERR rows blank — re-run to retry them")
    print(f"per-vote detail (flags, quotes, reasons) is in {JOURNAL.name}")


if __name__ == "__main__":
    main()
