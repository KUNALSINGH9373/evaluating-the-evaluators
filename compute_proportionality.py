#!/usr/bin/env python3
"""
compute_proportionality.py — Stage 9 for "Evaluating the Evaluators".

Proportionality is DERIVED, never hand-edited (rulebook §9 col 33 / methodology §12).
This script is the only thing that should ever write that column.

    Formula, no discretion, severity-dependent (v10, revised 2026-08-03).
    Computed on Channel A only, Tier A rows only. Policy uptake never substitutes.

                 Substantive   Partial              Acknowledged         None
        C1       Proportionate Under-response (gap) Under-response (gap) Accountability gap (no action)
        C2       Proportionate Proportionate        Under-response (gap) Accountability gap (no action)

    Equivalently: C1 needs a Substantive response to pass; C2 needs at least Partial.

No too-recent exception: rows carrying that modifier are already excluded upstream
at Action Trackable = no, so they never reach this formula.

Usage
-----
  python3 compute_proportionality.py            # dry run: report only, write nothing
  python3 compute_proportionality.py --write    # write the column and save

Coherence checks run either way, because §9 states them as invariants:
  * Action Trackable = yes  <=>  Action Level populated
  * Lag (days) exists       <=>  Response Date exists
  * Channel A Evidence exists <=> a response exists (level != None)
  * Proportionality populated <=> Tier A
"""

from __future__ import annotations

import argparse
import collections
import datetime
import shutil
import sys
import time
from pathlib import Path

import openpyxl

TARGET = Path.home() / "Desktop" / "v11_FINAL.xlsx"

SEV = "Severity (C1/C2) majority"
LEVEL = "Action Level"
PROP = "Proportionality"
TIER = "Action Trackable?"

MATRIX = {
    ("C1", "Substantive"): "Proportionate",
    ("C1", "Partial"): "Under-response (gap)",
    ("C1", "Acknowledged"): "Under-response (gap)",
    ("C1", "None"): "Accountability gap (no action)",
    ("C2", "Substantive"): "Proportionate",
    ("C2", "Partial"): "Proportionate",
    ("C2", "Acknowledged"): "Under-response (gap)",
    ("C2", "None"): "Accountability gap (no action)",
}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--target", type=Path, default=TARGET)
    ap.add_argument("--write", action="store_true", help="write the column (default is a dry run)")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.target)
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in ws[1]]
    ix = {h: i + 1 for i, h in enumerate(hdr) if h}
    for col in (SEV, LEVEL, PROP, TIER):
        if col not in ix:
            sys.exit(f"column missing from sheet: {col!r}")
    g = lambda r, k: ws.cell(r, ix[k]).value

    tier_a = [r for r in range(2, ws.max_row + 1) if g(r, TIER) == "yes"]
    print(f"{ws.max_row - 1} records · {len(tier_a)} Tier A")

    # --- blockers ------------------------------------------------------
    no_sev = [r for r in tier_a if g(r, SEV) not in ("C1", "C2")]
    no_lvl = [r for r in tier_a if not g(r, LEVEL)]
    if no_sev:
        print(f"\nBLOCKED: {len(no_sev)} Tier A rows have no C1/C2 severity majority.")
        print("  Run run_severity_ensemble.py --run first. Proportionality cannot be derived without it.")
        for r in no_sev[:5]:
            print(f"    {g(r, 'Finding ID')}  severity={g(r, SEV)!r}")
        if len(no_sev) > 5:
            print(f"    ... and {len(no_sev) - 5} more")
    if no_lvl:
        print(f"\nBLOCKED: {len(no_lvl)} Tier A rows have no Action Level.")
        for r in no_lvl[:5]:
            print(f"    {g(r, 'Finding ID')}")
    ready = [r for r in tier_a if r not in set(no_sev) | set(no_lvl)]
    print(f"\ncomputable now: {len(ready)}/{len(tier_a)}")

    # --- coherence (§9 invariants) --------------------------------------
    problems = []
    for r in range(2, ws.max_row + 1):
        fid = g(r, "Finding ID")
        is_a = g(r, TIER) == "yes"
        lvl = g(r, LEVEL)
        if is_a and not lvl:
            problems.append((fid, "Tier A but Action Level empty"))
        if lvl and not is_a:
            problems.append((fid, "Action Level set on a non-Tier-A row"))
        rd, lag = g(r, "Response Date"), g(r, "Lag (days)")
        if bool(rd) != (lag is not None and lag != ""):
            problems.append((fid, f"Response Date {rd!r} but Lag {lag!r}"))
        ev = g(r, "Channel A Evidence")
        if lvl and lvl != "None" and not ev:
            problems.append((fid, f"Action Level {lvl} but no Channel A Evidence"))
        if lvl == "None" and ev:
            problems.append((fid, "Action Level None but Channel A Evidence present"))
    print(f"coherence violations: {len(problems)}")
    for fid, msg in problems[:12]:
        print(f"    {fid}: {msg}")
    if len(problems) > 12:
        print(f"    ... and {len(problems) - 12} more")

    # --- derive ----------------------------------------------------------
    out = collections.Counter()
    cross = collections.Counter()
    for r in ready:
        key = (g(r, SEV), g(r, LEVEL))
        val = MATRIX.get(key)
        if val is None:
            print(f"  unmapped combination {key} on {g(r, 'Finding ID')} — skipped")
            continue
        cross[key] += 1
        out[val] += 1
        if a.write:
            ws.cell(r, ix[PROP]).value = val

    if ready:
        print("\nseverity x action level:")
        for sev in ("C1", "C2"):
            for lvl in ("Substantive", "Partial", "Acknowledged", "None"):
                n = cross.get((sev, lvl), 0)
                if n:
                    print(f"    {sev} + {lvl:<13} {n:>3}  -> {MATRIX[(sev, lvl)]}")
        print("\nProportionality:")
        tot = sum(out.values())
        for k in ("Proportionate", "Under-response (gap)", "Accountability gap (no action)"):
            n = out.get(k, 0)
            print(f"    {k:<32} {n:>3}  {n / tot:>6.1%}" if tot else f"    {k}  0")

    if a.write and ready:
        bak = a.target.with_name(a.target.stem + "-backup-" + time.strftime("%Y%m%d-%H%M%S") + a.target.suffix)
        shutil.copy2(a.target, bak)
        wb.save(a.target)
        print(f"\nwrote {len(ready)} rows to {a.target.name}  (backup: {bak.name})")
    elif not a.write:
        print("\ndry run — nothing written. Re-run with --write to commit.")


if __name__ == "__main__":
    main()
