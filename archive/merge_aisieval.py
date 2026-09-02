#!/usr/bin/env python3
"""
merge_aisieval.py — merge v10 revised + v11 final into a single AISIEVAL.xlsx.

Layout: one sheet, tier-major then date-ascending (rulebook tier order A -> B -> C,
publication date within tier).

The merge is not a concatenation. Three things are reconciled first, because the two
datasets were coded under rule revisions that landed between them:

  1. SCHEMA. v11 renamed 'Sources Checked' -> 'Sources Checked (channel A)' and dropped
     'Traceability Tag'. Aliases are mapped; v10-only columns are carried, not discarded.

  2. ATTRIBUTION (rulebook §9 col 18, revised 2026-08-15). Attribution is assessed only
     where a response exists; no response => 'Not applicable'. v10 predates this and codes
     those rows 'No explicit attribution'. Migrated here so the column means one thing
     across the merged file.

  3. PROPORTIONALITY (rulebook §9 col 33). Derived, never carried over. Recomputed from
     scratch for every Tier A row in both halves off the current matrix.

Finding ID is the merge key. A collision is reported, never silently resolved: v11 wins
as the revised half, and every collision is listed.

  python3 merge_aisieval.py                  # dry run: reconcile, verify, write nothing
  python3 merge_aisieval.py --write          # write ~/Desktop/AISIEVAL.xlsx
"""

from __future__ import annotations

import argparse
import collections
import csv
import datetime
import shutil
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DESKTOP = Path.home() / "Desktop"
V11 = DESKTOP / "v11_FINAL.xlsx"
V10_CANDIDATES = [
    Path.home() / "evaluating-the-evaluators" / "v10_revised.xlsx",
    DESKTOP / "v10 revised.xlsx",
    Path.home() / "evaluating-the-evaluators" / "v10_revised.csv",
]
OUT = DESKTOP / "AISIEVAL.xlsx"
SHEET = "AISIEVAL"

# v10 header -> v11 header. Headers are whitespace-stripped on both sides first,
# so 'Report Title ' and 'Report Title' collapse to one column rather than two.
ALIAS = {
    "Sources Checked": "Sources Checked (channel A)",
}

# Rulebook §9 col 25 names three Policy Level values. v11 carries a fourth spelling
# for 12 rows; it maps unambiguously onto the canonical one.
VALUE_FIX = {
    "Policy Level": {"Non-binding policy-related uptake": "Non-binding policy uptake"},
}

# Reported, never silently rewritten — an out-of-vocabulary value is a coding
# question for the researcher, not something a merge script should decide.
VOCAB = {
    "Severity (C1/C2) majority": {"C1", "C2"},
    "Action Level": {"None", "Acknowledged", "Partial", "Substantive"},
    "Attribution": {"Explicit attribution", "No explicit attribution", "Not applicable"},
    "Policy Level": {"No policy uptake identified", "Non-binding policy uptake", "Binding policy action"},
    "Proportionality": {"Proportionate", "Under-response (gap)", "Accountability gap (no action)"},
    "Eval? (trackable)": {"yes", "no"},
    "Action Trackable?": {"yes", "no"},
    "Scope": {"government-AISI", "third-party-evaluator"},
    "Access Type": {"Pre-deployment", "Post-deployment", "Mixed", "Aggregate", "N/A"},
    "Institution Type": {"Government", "For-Profit", "Non-Profit (AIEF)", "Non-Profit (Independent)", "Lab"},
    "Domain": {
        "Alignment", "Autonomy", "Bio-Chem", "Cyber", "Jailbreaks", "Societal", "Institutional",
        "Human Influence", "Eval-methodology", "Eval-tooling", "Frontier-forecasting",
        "Policy/Standards", "Transparency/Disclosure", "International-coordination",
    },
}
MULTIVALUED = {"Domain", "Institution Type"}

SEV, LEVEL, PROP, ATTR = "Severity (C1/C2) majority", "Action Level", "Proportionality", "Attribution"
EVAL, TRACK = "Eval? (trackable)", "Action Trackable?"
PROV = "Source Dataset"

MATRIX = {
    ("C1", "Substantive"): "Proportionate",
    ("C1", "Partial"): "Under-response (gap)",
    ("C1", "Acknowledged"): "Under-response (gap)",
    ("C1", "None"): "Accountability gap (no action)",
    ("C2", "Substantive"): "Proportionate",
    ("C2", "Partial"): "Proportionate",
    ("C2", "Acknowledged"): "Under-response (gap)",
    ("C2", "None"): "Accountability gap (no action)",
}

TIER_FILL = {"A": "FFFDECEC", "B": "FFFFF8E1", "C": "FFEFF6FF"}
HEADER_FILL = "FFD9D9D9"


def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def read_xlsx(path: Path) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [norm(c.value) for c in ws[1]]
    keep = [(i, h) for i, h in enumerate(hdr, 1) if h]
    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {h: norm(ws.cell(r, i).value) for i, h in keep}
        if any(rec.values()):
            rows.append(rec)
    return [h for _, h in keep], rows


def read_csv(path: Path) -> tuple[list[str], list[dict]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        rd = csv.DictReader(f)
        rows = [{norm(k): norm(v) for k, v in r.items() if k} for r in rd]
        return [norm(h) for h in rd.fieldnames if h], rows


REVIEW_COLS = ["Tier (candidate)", "Tier Changed?", "Tier Change Reason",
               "Classification Confidence", "Human Review"]


def read_v10_revised(path: Path):
    """'v10 revised' is an audit workbook, not a table.

    Sheet 'Original v10' holds the 40-column data; sheet 'Reclassification' holds the
    atomic tier audit whose formula column 'Derived Tier' supersedes the original
    Eval?/Action Trackable? pair. Data comes from the first, tiers from the second.
    One CoAI row was removed upstream, so 'Original v10' carries 455 live rows.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ov, rc = wb["Original v10"], wb["Reclassification"]

    oh = [norm(c.value) for c in ov[1]]
    rows = []
    for r in range(2, ov.max_row + 1):
        rec = {h: norm(ov.cell(r, i).value) for i, h in enumerate(oh, 1) if h}
        if rec.get("Finding ID"):
            rows.append(rec)

    rh = [norm(c.value) for c in rc[1]]
    rix = {h: i for i, h in enumerate(rh, 1) if h}
    tiers = {}
    for r in range(2, rc.max_row + 1):
        fid = norm(rc.cell(r, rix["Finding ID"]).value)
        if fid:
            tiers[fid] = {
                "original": norm(rc.cell(r, rix["Original Tier"]).value),
                "candidate": norm(rc.cell(r, rix["Derived Tier"]).value),
                "Tier Changed?": norm(rc.cell(r, rix["Tier Changed?"]).value),
                "Tier Change Reason": norm(rc.cell(r, rix["Accountability Exclusion Reason"]).value),
                "Classification Confidence": norm(rc.cell(r, rix["Classification Confidence"]).value),
                "Human Review": norm(rc.cell(r, rix["Human Review"]).value),
            }
    return [h for h in oh if h], rows, tiers


def load(path: Path):
    return read_csv(path) if path.suffix.lower() == ".csv" else read_xlsx(path)


def tier_of(rec) -> str:
    ev, tr = rec.get(EVAL, ""), rec.get(TRACK, "")
    if ev == "yes" and tr == "yes":
        return "A"
    if ev == "yes" and tr == "no":
        return "B"
    if ev == "no":
        return "C"
    return "?"


def datekey(rec) -> str:
    d = rec.get("Publication Date", "")
    return d if len(d) == 10 else (d + "-99-99")[:10]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--v10", type=Path)
    ap.add_argument("--v11", type=Path, default=V11)
    ap.add_argument("--out", type=Path, default=OUT)
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--tiers", choices=["original", "candidate"], default="original",
                    help="'original' (default) keeps v10's validated tiers live and carries the "
                         "atomic-rule proposal in review columns; 'candidate' applies the proposal. "
                         "The audit workbook states candidate tiers are not for headline estimates "
                         "until mandatory human adjudication is done.")
    a = ap.parse_args()

    v10_path = a.v10 or next((p for p in V10_CANDIDATES if p.exists()), None)
    if v10_path is None or not v10_path.exists():
        sys.exit("v10 revised not found. Tried:\n  " + "\n  ".join(str(p) for p in V10_CANDIDATES))

    h11, r11 = load(a.v11)
    revised = v10_path.suffix.lower() == ".xlsx" and "Reclassification" in openpyxl.load_workbook(v10_path).sheetnames
    if revised:
        h10, r10, tiers = read_v10_revised(v10_path)
    else:
        h10, r10 = load(v10_path)
        tiers = {}
    print(f"v11  {a.v11.name:<28} {len(r11):>4} records, {len(h11)} columns")
    print(f"v10  {v10_path.name:<28} {len(r10):>4} records, {len(h10)} columns")

    # --- 1. schema reconciliation ---------------------------------------
    r10 = [{ALIAS.get(k, k): v for k, v in rec.items()} for rec in r10]
    h10 = [ALIAS.get(h, h) for h in h10]
    # vestigial spacer, empty in both halves; already dropped from v11 during the repair
    DROP = {"(blank legacy column 34)"}
    h10 = [h for h in h10 if h not in DROP]
    for rec in r10:
        for c in DROP:
            rec.pop(c, None)

    only10 = [h for h in h10 if h not in h11]
    only11 = [h for h in h11 if h not in h10]
    print(f"\nschema · v10-only {only10 or 'none'} · v11-only {only11 or 'none'}")
    # Review columns only exist when v10 arrives as the audit workbook. Once the tiers
    # have been applied and the sheet flattened, they would be 1013 empty cells.
    columns = list(h11) + [h for h in only10 if h not in h11] + (REVIEW_COLS if tiers else []) + [PROV]

    for rec, tag in [(x, "v11") for x in r11] + [(x, "v10") for x in r10]:
        rec[PROV] = tag

    # --- 1a. apply the atomic tier reclassification to the v10 half -------
    # Tier-A-scoped judgement columns. Evidence columns (Company Response, verbatims,
    # dates, lag) are never cleared — the evidence stays true whatever the tier.
    A_ONLY = [LEVEL, ATTR, "Policy Level", PROP]
    promoted, demoted, cleared = [], [], 0
    if tiers:
        for rec in r10:
            t = tiers.get(rec["Finding ID"])
            if not t:
                continue
            for c in REVIEW_COLS[1:]:
                rec[c] = t[c]
            rec["Tier (candidate)"] = t["candidate"]
            if a.tiers == "candidate":
                rec[EVAL], rec[TRACK] = {"A": ("yes", "yes"), "B": ("yes", "no"),
                                         "C": ("no", "")}[t["candidate"]]
                if t["original"] == t["candidate"]:
                    continue
                if t["candidate"] == "A":
                    promoted.append(rec["Finding ID"])
                elif t["original"] == "A":
                    demoted.append(rec["Finding ID"])
                    kept = [f"{c}={rec[c]}" for c in A_ONLY if rec.get(c)]
                    if kept:
                        rec["Notes"] = (rec.get("Notes", "") + " " + f"[tier revised A->{t['candidate']}"
                                        f" 2026-08-15; cleared Tier-A-only codings: {'; '.join(kept)}]").strip()
                        cleared += 1
                    for c in A_ONLY:
                        rec[c] = ""
        n_chg = sum(1 for t in tiers.values() if t["Tier Changed?"] == "Yes")
        n_mand = sum(1 for t in tiers.values() if t["Human Review"] == "Mandatory")
        if a.tiers == "candidate":
            print(f"tier rule · CANDIDATE tiers APPLIED to {len(tiers)} v10 rows")
            print(f"    promoted to Tier A: {len(promoted)} · demoted: {len(demoted)} ({cleared} codings moved to Notes)")
        else:
            print(f"tier rule · validated v10 tiers kept live; candidate tiers carried in review columns")
            print(f"    {n_chg} rows the audit proposes changing · {n_mand} flagged Mandatory human review")

    # --- 1b. controlled-vocabulary normalisation --------------------------
    fixed = collections.Counter()
    for rec in r10 + r11:
        for col, mapping in VALUE_FIX.items():
            if rec.get(col) in mapping:
                fixed[(rec[PROV], col, rec[col], mapping[rec[col]])] += 1
                rec[col] = mapping[rec[col]]
    print(f"vocabulary · normalised {sum(fixed.values())} values")
    for (src, col, old, new), n in fixed.most_common():
        print(f"    {src} {col}: {old!r} -> {new!r}  x{n}")

    # --- 2. attribution migration (v10 half) -----------------------------
    mig = 0
    for rec in r10:
        if tier_of(rec) == "A" and rec.get(LEVEL) == "None" and rec.get(ATTR) != "Not applicable":
            rec[ATTR] = "Not applicable"
            mig += 1
    print(f"attribution · migrated {mig} v10 rows to 'Not applicable' (no response)")

    # --- 3. merge on Finding ID -------------------------------------------
    by_id, collisions = {}, []
    for rec in r11:
        by_id[rec["Finding ID"]] = rec
    for rec in r10:
        fid = rec["Finding ID"]
        if fid in by_id:
            diff = [c for c in columns if c != PROV and norm(rec.get(c)) != norm(by_id[fid].get(c))]
            collisions.append((fid, diff))
        else:
            by_id[fid] = rec
    print(f"merge · {len(by_id)} unique Finding IDs · {len(collisions)} collisions (v11 kept)")
    for fid, diff in collisions[:15]:
        print(f"    {fid}: {len(diff)} differing columns {diff[:5]}")
    if len(collisions) > 15:
        print(f"    ... and {len(collisions) - 15} more")

    merged = list(by_id.values())

    # --- 4. proportionality: recompute, never carry ------------------------
    changed, pending = collections.Counter(), []
    for rec in merged:
        if tier_of(rec) == "A":
            if not rec.get(LEVEL):
                # Newly promoted by the atomic rule; Channel A not yet coded.
                pending.append(rec["Finding ID"])
                rec[PROP] = ""
                continue
            new = MATRIX.get((rec.get(SEV), rec.get(LEVEL)))
            if new is None:
                print(f"    unmapped {rec['Finding ID']}: sev={rec.get(SEV)!r} AL={rec.get(LEVEL)!r}")
                continue
            if rec.get(PROP) != new:
                changed[(rec[PROV], rec.get(PROP) or "(blank)", new)] += 1
            rec[PROP] = new
        else:
            rec[PROP] = ""
    print(f"proportionality · recomputed · {sum(changed.values())} values changed")
    for (src, old, new), n in changed.most_common():
        print(f"    {src}: {old!r} -> {new!r}  x{n}")
    if pending:
        print(f"    {len(pending)} Tier A rows left blank — Channel A not yet coded (see report)")

    # --- 5. order: tier-major, date-ascending -----------------------------
    order = {"A": 0, "B": 1, "C": 2, "?": 3}
    merged.sort(key=lambda r: (order[tier_of(r)], datekey(r), r["Finding ID"]))

    # --- 6. verification ---------------------------------------------------
    print("\n--- verification ---")
    ids = [r["Finding ID"] for r in merged]
    dup = [k for k, v in collections.Counter(ids).items() if v > 1]
    exp = len(r11) + len(r10) - len(collisions)
    print(f"records {len(merged)} (expected {exp}) · duplicates {dup or 'none'} · blank IDs {sum(1 for i in ids if not i)}")
    assert len(merged) == exp and not dup and all(ids)

    lost = [fid for fid in {r["Finding ID"] for r in r11} | {r["Finding ID"] for r in r10} if fid not in set(ids)]
    print(f"records dropped from either input: {lost or 'none'}")
    assert not lost

    tiers = collections.Counter(tier_of(r) for r in merged)
    print("tiers  " + " · ".join(f"{k} {tiers[k]}" for k in "ABC?" if tiers[k]))
    assert tiers["?"] == 0, "rows with an unrecognised tier"

    prov = collections.Counter(r[PROV] for r in merged)
    print("source " + " · ".join(f"{k} {v}" for k, v in sorted(prov.items())))
    for col in (SEV, LEVEL, ATTR, PROP, "Policy Level"):
        c = collections.Counter(r.get(col) for r in merged if r.get(col))
        print(f"{col:<26} " + " · ".join(f"{k} {v}" for k, v in c.most_common()))

    oov = collections.Counter()
    for r in merged:
        for col, allowed in VOCAB.items():
            v = r.get(col, "")
            if not v:
                continue
            parts = [p.strip() for p in v.split(";")] if col in MULTIVALUED else [v]
            for p in parts:
                if p not in allowed:
                    oov[(col, p, r[PROV])] += 1
    print(f"out-of-vocabulary values: {sum(oov.values())} across {len(oov)} distinct")
    for (col, val, src), n in sorted(oov.items(), key=lambda x: -x[1]):
        print(f"    {col:<26} {val!r:<44} {src}  x{n}")

    bad = []
    for r in merged:
        t = tier_of(r)
        if t == "A" and not r.get(LEVEL):
            bad.append((r["Finding ID"], "Tier A without Action Level"))
        if t != "A" and r.get(LEVEL):
            bad.append((r["Finding ID"], "Action Level on non-Tier-A row"))
        if t != "A" and r.get(PROP):
            bad.append((r["Finding ID"], "Proportionality on non-Tier-A row"))
        if r.get(LEVEL) == "None" and r.get(ATTR) != "Not applicable":
            bad.append((r["Finding ID"], f"no response but Attribution {r.get(ATTR)!r}"))
        if bool(r.get("Response Date")) != bool(r.get("Lag (days)") not in ("", None)):
            bad.append((r["Finding ID"], "Response Date / Lag mismatch"))
    print(f"coherence violations: {len(bad)}")
    for fid, m in bad[:15]:
        print(f"    {fid}: {m}")

    if not a.write:
        print("\ndry run — nothing written. Re-run with --write.")
        return

    # --- 7. write -----------------------------------------------------------
    if a.out.exists():
        shutil.copy2(a.out, a.out.with_name(a.out.stem + ".backup-" + time.strftime("%Y%m%d-%H%M%S") + a.out.suffix))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(columns)
    for rec in merged:
        ws.append([rec.get(c, "") for c in columns])

    wide = {"Finding ID": 30, "Report ID": 26, "Institution": 30, "Institution Type": 18,
            "Report Title ": 44, "Source URL": 44, "Finding": 70, "Finding Quote": 70,
            "Models / Systems": 30, "Notes": 44}
    for i, c in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = wide.get(c, 18)
        cell = ws.cell(1, i)
        cell.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for j, rec in enumerate(merged, start=2):
        rgb = TIER_FILL[tier_of(rec)]
        fill = PatternFill("solid", start_color=rgb, end_color=rgb)
        for i in range(1, len(columns) + 1):
            ws.cell(j, i).fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
    wb.save(a.out)

    chk = openpyxl.load_workbook(a.out)[SHEET]
    print(f"\nwrote {a.out}")
    print(f"  {chk.max_row - 1} records × {chk.max_column} columns · sheet {SHEET!r}")


if __name__ == "__main__":
    main()
