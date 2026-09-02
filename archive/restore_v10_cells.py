#!/usr/bin/env python3
"""
restore_v10_cells.py — put back the evidence cells that the v10 flattening dropped.

Provenance established by diffing three snapshots of the same 455 findings:

    v10.csv  ==  audit workbook 'Original v10'  !=  flattened 'v10 revised'

The audit workbook is a byte-perfect copy of v10.csv on every column, so the
reclassification lost nothing. The loss happened when that workbook was flattened
into a single sheet: ~650 evidence cells came out empty, well beyond the ~22 cells
the demotions legitimately clear.

Restore rule: a cell is refilled only when it is EMPTY in the flattened file and
POPULATED in v10.csv. Nothing already present is touched, so every revision in
'v10 revised' — the atomic tiers, the Attribution migration, the demotion clears —
survives untouched.

Three exclusions, because these were emptied on purpose:
  * Tier-A-only judgement columns (Action Level, Attribution, Policy Level,
    Proportionality) are never restored to a row that is no longer Tier A.
  * 'Action Trackable?' / 'Eval? (trackable)' are never restored — the atomic
    reclassification is authoritative for tier, including blank on Tier C.
  * Proportionality is not restored at all; it is recomputed from the matrix.

  python3 restore_v10_cells.py            # dry run
  python3 restore_v10_cells.py --write    # write ~/Desktop/v10_revised_RESTORED.xlsx
"""

from __future__ import annotations

import argparse
import collections
import csv
import datetime
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC_CSV = Path.home() / "evaluating-the-evaluators" / "v10.csv"
FLAT = Path.home() / "Documents" / "v10_revised.xlsx"
OUT = FLAT                      # restored in place; a timestamped backup is taken first

ALIAS = {"Sources Checked": "Sources Checked (channel A)"}
A_ONLY = {"Action Level", "Attribution", "Policy Level", "Proportionality"}
NEVER = {"Eval? (trackable)", "Action Trackable?", "Proportionality"}

MATRIX = {
    ("C1", "Substantive"): "Proportionate",
    ("C1", "Partial"): "Under-response (gap)",
    ("C1", "Acknowledged"): "Under-response (gap)",
    ("C1", "None"): "Accountability gap (no action)",
    ("C2", "Substantive"): "Proportionate",
    ("C2", "Partial"): "Proportionate",
    ("C2", "Acknowledged"): "Under-response (gap)",
    ("C2", "None"): "Accountability gap (no action)",
}
TIER_FILL = {"A": "FFFDECEC", "B": "FFFFF8E1", "C": "FFEFF6FF"}


def norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--out", type=Path, default=OUT)
    a = ap.parse_args()

    old = {}
    with open(SRC_CSV, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rec = {ALIAS.get(k.strip(), k.strip()): norm(v) for k, v in r.items() if k}
            old[rec["Finding ID"]] = rec

    wb = openpyxl.load_workbook(FLAT, data_only=True)
    ws = wb["v10 revised"]
    hdr = [norm(c.value) for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {h: norm(ws.cell(r, i).value) for i, h in enumerate(hdr, 1) if h}
        if rec.get("Finding ID"):
            rows.append(rec)
    print(f"v10.csv {len(old)} rows · flattened {len(rows)} rows · {len(hdr)} columns")

    tier = lambda rec: ("A" if rec.get("Action Trackable?") == "yes"
                        else "B" if rec.get("Eval? (trackable)") == "yes" else "C")

    restored = collections.Counter()
    skipped = collections.Counter()
    per_row = collections.Counter()
    missing = [f for f in (r["Finding ID"] for r in rows) if f not in old]
    if missing:
        print(f"  !! {len(missing)} rows absent from v10.csv, left as-is: {missing[:5]}")

    for rec in rows:
        o = old.get(rec["Finding ID"])
        if not o:
            continue
        t = tier(rec)
        for col in hdr:
            if not col or col in NEVER:
                continue
            if rec.get(col) or not o.get(col):
                continue
            if col in A_ONLY and t != "A":
                skipped[col] += 1
                continue
            rec[col] = o[col]
            restored[col] += 1
            per_row[rec["Finding ID"]] += 1

    print(f"\nrestored {sum(restored.values())} cells across {len(per_row)} rows")
    for c, n in restored.most_common():
        print(f"    {c:<32} {n:>4}")
    if skipped:
        print("\nwithheld (Tier-A-only column on a demoted row):")
        for c, n in skipped.most_common():
            print(f"    {c:<32} {n:>4}")

    # Lag is derivable; fill any that are still blank but have a Response Date.
    derived = 0
    for rec in rows:
        if rec.get("Response Date") and not rec.get("Lag (days)"):
            try:
                d0 = datetime.date.fromisoformat(rec["Publication Date"][:10])
                d1 = datetime.date.fromisoformat(rec["Response Date"][:10])
                rec["Lag (days)"] = str((d1 - d0).days)
                derived += 1
            except ValueError:
                pass
    print(f"\nLag (days) derived from Response Date - Publication Date: {derived}")

    # Proportionality: always recomputed, never carried.
    prop = collections.Counter()
    for rec in rows:
        if tier(rec) == "A" and rec.get("Action Level"):
            v = MATRIX.get((rec.get("Severity (C1/C2) majority"), rec["Action Level"]))
            if v:
                prop[v] += 1
                rec["Proportionality"] = v
        else:
            rec["Proportionality"] = ""
    print("Proportionality recomputed: " + " · ".join(f"{k} {v}" for k, v in prop.most_common()))

    # --- verification ----------------------------------------------------
    print("\n--- verification ---")
    tiers = collections.Counter(tier(r) for r in rows)
    print("tiers  " + " · ".join(f"{k} {tiers[k]}" for k in "ABC"))
    A = [r for r in rows if tier(r) == "A"]
    print(f"Tier A {len(A)} · missing Action Level {sum(1 for r in A if not r.get('Action Level'))}"
          f" · missing Policy Level {sum(1 for r in A if not r.get('Policy Level'))}")
    bad = sum(1 for r in rows if tier(r) != "A" and any(r.get(c) for c in A_ONLY))
    print(f"Tier-A-only columns on non-Tier-A rows: {bad}")
    lag = sum(1 for r in rows if bool(r.get("Response Date")) != bool(r.get("Lag (days)")))
    print(f"Response Date / Lag mismatches: {lag}")
    still = sum(1 for r in rows for c in hdr
                if c and c not in NEVER and not r.get(c) and old.get(r["Finding ID"], {}).get(c)
                and not (c in A_ONLY and tier(r) != "A"))
    print(f"cells still empty that v10.csv has (should be 0): {still}")

    if not a.write:
        print("\ndry run — nothing written.")
        return

    if a.out.exists():
        import shutil, time
        bak = a.out.with_name(a.out.stem + ".backup-" + time.strftime("%Y%m%d-%H%M%S") + a.out.suffix)
        shutil.copy2(a.out, bak)
        print(f"\nbackup: {bak}")

    out = openpyxl.Workbook()
    o = out.active
    o.title = "v10 revised"
    o.append(hdr)
    for rec in rows:
        o.append([rec.get(c, "") for c in hdr])
    wide = {"Finding ID": 30, "Report ID": 26, "Institution": 30, "Report Title": 44,
            "Source URL": 44, "Finding": 70, "Finding Quote": 70, "Models / Systems": 30, "Notes": 44}
    for i, c in enumerate(hdr, 1):
        o.column_dimensions[get_column_letter(i)].width = wide.get(c, 18)
        cell = o.cell(1, i)
        cell.fill = PatternFill("solid", start_color="FFD9D9D9", end_color="FFD9D9D9")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for j, rec in enumerate(rows, start=2):
        rgb = TIER_FILL[tier(rec)]
        fill = PatternFill("solid", start_color=rgb, end_color=rgb)
        for i in range(1, len(hdr) + 1):
            o.cell(j, i).fill = fill
    o.freeze_panes = "A2"
    o.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{o.max_row}"
    out.save(a.out)
    print(f"\nwrote {a.out}  ({o.max_row - 1} records × {o.max_column} columns)")


if __name__ == "__main__":
    main()
